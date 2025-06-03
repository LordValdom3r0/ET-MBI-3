"""
        NOTAS:

for root, dirs, files in os.walk(dir_path):
    for file in files: 
  
        # change the extension from '.mp3' to 
        # the one of your choice.
        if file.endswith('.mp3'):
            print (root+'/'+str(file))

#Para la caja TBLU se usan fusibles ATO con color claro por lo que al color se le agrega un "_clear", por ejemplo "ATO,10,red_clear"

#"F400": "ATO,15,BLUE
"""
           
from copy import copy
import requests
import openpyxl
import json
import os
from model import model
import gc
modules = {}
modules_t = {}

datos_conexion=model()
host,user,password,database,serverp2,dbp2,userp2,passwordp2,printerhost=datos_conexion.datos_acceso()

fuses_color = {
    "1":    "black",
    "5":    "beige",
    "7.5":  "cafe",
    "10":   "rojo",
    "15":   "azul",
    "20":   "amarillo",
    "25":   "natural",
    "30":   "verde",
    "40":   "naranja",
    "50":   "rojo",
    "60":   "azul"
    }

torques_value = {
  'PDC-P': {'E1': ""},
  'PDC-D': {'E1': ""},
  'PDC-R': {'E1': ""},
  'PDC-RMID': {'E1': ""},
  'PDC-RS': {'E1': ""},
  "MFB-P1": {"A41": "", "A42": "", "A43": "", "A44": "", "A45": "", "A46": "","A47": "",}, 
  "MFB-S": {"A51": "","A52": "", "A53": "", "A54": "", "A55": "", "A56": ""},
  'MFB-E': {"E1": "", "A1": "", "A2": ""},
  "MFB-P2": {"A20": "","A21": "", "A22": "", "A23": "", "A24": "", "A25": "", "A26": "", "A27": "", "A28": "", "A29": "", "A30": ""},
  "BATTERY": {"BT": ""},
  "BATTERY-2": {"BT": ""},
  "BATTERY-3": {"BT": ""}
}

##################################### Modules management #################################
def makeModules(data):
    global modules, modules_t
    #print("Data dentro de la creación de módulos!: ",data)
    # Se manda llamar a la función encargada de consultar los módulos determinantes desde la base de datos, para posteriormente meterlos en un json llamado "pdcrVariantes".
    endpoint = f"http://{host}:5000/api/get/{data}/pdcr/variantes"
    pdcrVariantes = requests.get(endpoint).json()
    print("Lista Final de Variantes PDC-R: \n",pdcrVariantes)
    modules = {}
    modules_t = {}
    print("#################### Modules ####################")
    print("Modulos actual en MODULES: ",modules)
    print("Modulos actual en MODULEST: ",modules_t)
    dir_path = os.path.join(os.getcwd(), '..\\modules\\')
    file_name = None
    for root, dirs, files in os.walk(dir_path):
        for file_name in files: 
            print("file_name: ",file_name)
            if file_name.endswith('.xls') or file_name.endswith('.xlsx'):
                file = openpyxl.load_workbook(filename = dir_path + file_name, data_only=True)
                sheets = file.sheetnames
                for sheet in sheets: #para skipear las sheet que son correspondientes a información o X294 Izquierda F96
                    if "Acomodos Modularidades" in sheet or "X294 Izquierda F96" in sheet:
                        continue
                    #para torques
                    if "MFB" in sheet or "BATTERY" in sheet:
                        currentSheet = file[sheet]
                        for column in range(11, currentSheet.max_column + 1):
                            module = currentSheet.cell(row = 3, column = column).value

                            if isinstance(module,str):
                                module = module.replace(" ","")#se eliminan posibles espacios existentes, solo en str, porque puede haber valores None
                            print("Modulo: ",module)

                            if not(module in modules_t):
                                modules_t[module] = {}
                            for row in range(5,currentSheet.max_row  + 1):
                                value = currentSheet.cell(row = row, column = column).value

                                if isinstance(value,str):
                                    value = value.replace(" ","")#se eliminan posibles espacios existentes, si hay datos de lo contrario es None

                                if value == "x" or value == "X":
                                    box = currentSheet.cell(row = row, column = 1).value
                                    if isinstance(box,str):
                                        box = box.replace(" ","")#se eliminan posibles espacios existentes, si hay datos de lo contrario es None

                                    if box == "MFB-E":
                                        print("Modulo de Nueva Caja: ",module)
                                    torque = currentSheet.cell(row = row, column = 2).value
                                    if isinstance(torque,str):
                                        torque = torque.replace(" ","")#se eliminan posibles espacios existentes, si hay datos de lo contrario es None

                                    if box == "MFB-S" and torque== "G1/21":
                                        print("Aquí viene una battery-2",module)
                                        box = "BATTERY-2"
                                        torque = "BT"
                                        print("Box Actualizado: ",box)
                                        print("Torque Actualizado: ",torque)   
                                    if box == "MFB-S2" and torque== "G1/21":
                                        print("Aquí viene una battery-3",module,' : ',box)
                                        box = "BATTERY-3"
                                        torque = "BT"
                                        print("Box Actualizado: ",box)
                                        print("Torque Actualizado: ",torque)
                                    if box == "MFB-S2" and torque!= "G1/21":
                                        print("Aquí viene una MFB-S2",module,' : ',box)
                                        box ="MFB-P1"
                                        print("Caja Actualizada: ",box)
                                    if not(box in modules_t[module]):
                                        modules_t[module][box] = {}
                                    modules_t[module][box][torque] = True
                    #para fusibles
                    else:
                        currentSheet = file[sheet]
                        for column in range(8, currentSheet.max_column + 1):
                            module = currentSheet.cell(row = 3, column = column).value #se obtiene el valor de la celda que contiene el nombre del módulo

                            if isinstance(module,str):
                                module = module.replace(" ","")#se eliminan posibles espacios existentes, solo en str, porque puede haber valores None
                            print("Modulo: ",module)

                            if not(module in modules):
                                modules[module] = {}
                            for row in range(5,currentSheet.max_row  + 1):
                                value = currentSheet.cell(row = row, column = column).value #se obtiene el valor de la celda ej: "X"

                                if isinstance(value,str):
                                    value = value.replace(" ","")#se eliminan posibles espacios existentes, si hay datos de lo contrario es None

                                if value == "x" or value == "X":
                                    box = currentSheet.cell(row = row, column = 1).value.strip() #se obtiene el nombre de la caja primer columna de excel ej: PDC-P

                                    if isinstance(box,str):
                                        box = box.replace(" ","")#se eliminan posibles espacios existentes, si hay datos de lo contrario es None

                                    if box =="Fuse Box F55" or box =="FuseBoxF55":
                                        box = "TBLU"
                                    fuse = currentSheet.cell(row = row, column = 2).value #ejemplo: F210
                                    if isinstance(fuse,str):
                                        fuse = fuse.replace(" ","")#se eliminan posibles espacios existentes, si hay datos de lo contrario es None

                                    if box == "TBLU":
                                        fuse = fuse.replace("A", "")
                                    if "PDC-S" in box:
                                        fuse = str(fuse)
                                        #print("Tipo del Fuse Ya convertido: ",type(fuse))
                                    if "F96" in box:
                                        print("Caja F96 AQUI",module)
                                    if box == "PDC-R":
                                        if module in pdcrVariantes["large"]:
                                            box = "PDC-R"
                                        elif module in pdcrVariantes["medium"]:
                                            box = "PDC-RMID"
                                        elif module in pdcrVariantes["small"]:
                                            box = "PDC-RS"
                                        else:
                                            box = "PDC-RS"
                                        #print("VARIANTE PARA PDC-R DEL MÓDULO: ",box)
                                        if fuse == "X" or fuse == "T" or fuse == "U":
                                            fuse = "REL" + fuse
                                    amp = currentSheet.cell(row = row, column = 7).value #se obtiene el valor de la celda, ej: 7.5A

                                    if isinstance(amp,str):
                                        amp = amp.replace(" ","")#se eliminan posibles espacios existentes, si hay datos de lo contrario es None
                                    elif isinstance(amp,int):
                                        amp = str(amp) + "A" #es int porque en el excel le faltó la letra A de amperes
                                        print("____________________________________________")
                                        print("revisar excel, se agregó una A en: ",amp)
                                        print("____________________________________________")
                                    elif isinstance(amp,float):
                                        amp = str(amp) + "A" #es un float porque en el excel le faltó la letra A
                                        print("____________________________________________")
                                        print("revisar excel, se agregó una A en: ",amp)
                                        print("____________________________________________")

                                    if not(box in modules[module]):
                                        modules[module][box] = {}
                                    modules[module][box][fuse] = amp[:-1]
                del file
                gc.collect()
                os.remove(root+'\\'+ file_name)

    structured_data = []
    for module in modules:
        temp = {
            "DBEVENT": data,
            "MODULO": "",
            "CAJA_1": {},
            "CAJA_2": {},
            "CAJA_3": {},
            "CAJA_4": {},
            "CAJA_5": {},
            "CAJA_6": {},
            "CAJA_7": {},
            "CAJA_8": {},
            "CAJA_9": {},
            "CAJA_10": {},
            "CAJA_11": {},
            "CAJA_12": {},
            "CAJA_13": {},
            "CAJA_14": {},
            "CAJA_15": {},
            "CAJA_16": {}
            }

        temp["MODULO"] = module
        cnt = 1
        for box in modules[module]:
            try:
                if "PDC-P" in box:
                    if not box in modules_t[module]:
                        modules_t[module][box] = {"E1": True}
                if "PDC-D" in box:
                    if not box in modules_t[module]:
                        modules_t[module][box] = {"E1": True}
                if "PDC-R" in box:
                    if not box in modules_t[module]:
                        modules_t[module][box] = {"E1": True}
            except Exception as ex:
                print (ex)
                #print("Inconcistencia de modulos:")
                #print("Caja: ", box)
                #print("Modulo de vision: ",module)
                #print("Modulo inexistente para torque")
            key = "CAJA_" + str(cnt)
            if not(box in temp[key]):
                temp[key][box]  = {}
            for fuse in modules[module][box]:
                try:
                    amp     = modules[module][box][fuse]
                    color = ""
                    if amp == "60":
                        color = "1008695"
                    elif amp == "70":
                        color = "1010733"
                    else:
                        color   = fuses_color[amp]
                    temp[key][box][fuse] = color
                except Exception as ex:
                    print("\n Vision exception in [", module, "] [", box, "] [", fuse, "]")
                    print(ex)
            cnt += 1
        structured_data.append(temp)

    print ("\n total de modulos vision: ",len(structured_data))

    #print("modules_t:")
    #print(modules_t)


    torque_data = []
    for module in modules_t:
        temp = {
            "DBEVENT": data,
            "MODULO": "",
            "CAJA_1": {},
            "CAJA_2": {},
            "CAJA_3": {},
            "CAJA_4": {},
            "CAJA_5": {},
            "CAJA_6": {},
            "CAJA_7": {},
            "CAJA_8": {},
            "CAJA_9": {},
            "CAJA_10": {},
            "CAJA_11": {},
            "CAJA_12": {},
            "CAJA_13": {},
            "CAJA_14": {},
            "CAJA_15": {},
            "CAJA_16": {}
            }

        temp["MODULO"] = module
        cnt = 1
        for box in modules_t[module]:
            key = "CAJA_" + str(cnt)
            if not(box in temp[key]):
                temp[key][box]  = {}
            try:
                for torque in modules_t[module][box]:
                    try:
                        temp[key][box][torque] = modules_t[module][box][torque]
                    except Exception as ex:
                        print("\nTorque exception in [", module, "] [", box, "] [", fuse, "]")
                        print(ex)
            except Exception as ex:
                print(ex)
            cnt += 1
        key = "CAJA_" + str(cnt)
        if not("BATTERY" in temp[key]):
            temp[key]["BATTERY"]  = {"BT": True}
        torque_data.append(temp)

    print ("\n total de modulos de torque: ",len(torque_data))

    return structured_data, torque_data

def visionUpdate(data):
    print("vision updating")
    tabla = data[0]["DBEVENT"]
    #print("TABLAAAAA Vision: ",tabla)
    endpoint = f"http://{host}:5000/api/get/{tabla}/modulos_fusibles/all/-/-/-/-/-"
    existing = requests.get(endpoint).json()
    if not("MODULO" in existing):
        existing["MODULO"] = []
    for i in data:
        try:
            if not(i["MODULO"] in existing["MODULO"]):
                endpoint = f"http://{host}:5000/api/post/modulos_fusibles"
                response = requests.post(endpoint, data = json.dumps(i))
            else:
                #pass
                index = existing["MODULO"].index(i["MODULO"])
                id = existing["ID"][index]
                endpoint = f"http://{host}:5000/api/update/modulos_fusibles/{id}"
                response = requests.post(endpoint, data = json.dumps(i))
        except Exception as ex:
            print (ex)

def torqueUpdate(data):
    print("torque updating")
    tabla = data[0]["DBEVENT"]
    #print("TABLAAAAA Torque: ",tabla)
    endpoint = f"http://{host}:5000/api/get/{tabla}/modulos_torques/all/-/-/-/-/-"
    existing = requests.get(endpoint).json()
    if not("MODULO" in existing):
        existing["MODULO"] = []
    for i in data:
        try:
            if not(i["MODULO"] in existing["MODULO"]):
                endpoint = f"http://{host}:5000/api/post/modulos_torques"
                response = requests.post(endpoint, data = json.dumps(i))
            else:
                #pass
                index = existing["MODULO"].index(i["MODULO"])
                id = existing["ID"][index]
                endpoint = f"http://{host}:5000/api/update/modulos_torques/{id}"
                response = requests.post(endpoint, data = json.dumps(i))
        except Exception as ex:
            print (ex)

def pdcrVariants (data):
    """

            IN CONSTRUCTION


    PDC-R small:  A2239060902
    PDC-R MEDIUM:  A2239061002
    PDC-R LARGE:  A2239061102
    """
    print("#################### pdcrVariants ####################")
    dir_path = os.path.join(os.getcwd(), '..\\FAAJISPREV\\')
    file_name = None
    rows = []
    for root, dirs, files in os.walk(dir_path):
        for file_name in files: 
            temp = file_name.lower()
            ILX = temp.split(sep = ".")[0].upper()
            if temp.endswith('.txt'):
                fic = open(dir_path + file_name)
                lines = list(fic)
                for i in lines:
                    i = i[:-1]
                    rows.append(i.split())
                print(len(lines))
                for i in range(5):
                    print(lines[i])

def refreshModules(data):
    vision, torque = makeModules(data)
    visionUpdate(vision)
    torqueUpdate(torque)

################################### Modularities management ##############################
def makeModularities(data):
    flujo = ""
    numero = ""
    if 'izquierda' in data:
        print('EVENTO DE CONDUCCION IZQUIERDA')
        if 'z296' in data or 'Z296' in data:
            flujo = 'ILZ'
            numero = '296'
        if 'x296' in data or 'X296' in data:
            flujo = 'ILX'
            numero = '296'
        if 'x294' in data or 'X294' in data: 
            flujo = 'ILX'
            numero = '294'
    if 'derecha' in data:
        print('EVENTO DE CONDUCCION DERECHA')
        if 'z296' in data or 'Z296' in data:
            flujo = 'IRZ'
            numero = '296'
        if 'x296' in data or 'X296' in data:
            flujo = 'IRX'
            numero = '296'
        if 'x294' in data or 'X294' in data: 
            flujo = 'IRX'
            numero = '294'
            
    flujo_numero = flujo + numero


    global modules
    print("Dentro de MakeModularities DATA: ",data)
    # Se manda llamar a la función encargada de consultar los módulos determinantes desde la base de datos, para posteriormente meterlos en un json llamado "pdcrVariantes".
    endpoint = f"http://{host}:5000/api/get/{data}/pdcr/variantes"
    pdcrVariantes = requests.get(endpoint).json()
    print("Lista Final de Variantes PDC-R:\n",pdcrVariantes)
    print("#################### Modularities ####################")
    endpoint = f"http://{host}:5000/api/get/{data}/modulos_fusibles/all/-/-/-/-/-"
    modulesExisting = requests.get(endpoint).json()
    #print("Modulos existentes en la base de datos VISION: ",modulesExisting["MODULO"])
    print("LEN VISION: ",len(modulesExisting["MODULO"]))

    endpoint = f"http://{host}:5000/api/get/{data}/modulos_torques/all/-/-/-/-/-"
    modulesExisting_t = requests.get(endpoint).json()
    #print("Modulos existentes en la base de datos TORQUES: ",modulesExisting_t["MODULO"])
    print("LEN TORQUES: ",len(modulesExisting_t["MODULO"]))

    dir_path = os.path.join(os.getcwd(), '..\\ILX\\')
    file_name = None
    modularities = []
    modulosFaltantes = []
    ilxfaltantes = {
        "ILX": {},
        "Modulos": {}
        }
    


    for root, dirs, files in os.walk(dir_path):
        for file_name in files: 
            temp = file_name.lower()
            ILX = temp.split(sep = ".")[0].upper()
            if not(flujo_numero in file_name):# SI NO se encuentra el nombre esperado de inicio para un arnés de este tipo:
                ilxfaltantes["ILX"][ILX] = {
                            "vision": [],
                            "torque": []
                            } #se crea un diccionario para esta modularidad
                ilxfaltantes["ILX"][ILX]["vision"].append("No es un DAT válido para este evento") #se crea un diccionario para esta modularidad
                ilxfaltantes["ILX"][ILX]["torque"].append("No es un DAT válido para este evento") #se agrega el mensaje que no es un DAT válido
                modulosFaltantes.append(ILX) #se agrega a la lista final de módulos faltantes para que aparezca en pantalla
                ilxfaltantes["Modulos"] = modulosFaltantes #se actualiza esta lista
                os.remove(root+'\\'+ file_name) #se elimina el archivo de los DATS
            else:
                if temp.endswith('.dat'):
                    flag_s = False
                    flag_m = False
                    flag_l = False
                    qr_pdcr = {}
                    flag_mfbp2_der = False
                    flag_mfbp2 = []
                    fic = open(dir_path + file_name)
                    lines = list(fic)
                    csv = ""
                    for line in lines:
                        csv += line.rsplit(sep = "=")[-1][:-1] + ","
                    csv = csv[:-1]
                    fic.close()
                    #print("MODULOS DEL ILX: ",csv.split(sep = ","))
                    if "ILX294" in ILX:
                        print("Evento 294 IZQUIERDA")
                    if "IRX294" in ILX:
                        print("Evento 294 DERECHA")
                    if "ILX296" in ILX:
                        print("Evento 296 IZQUIERDA")
                    if "IRX296" in ILX:
                        print("Evento 296 DERECHA")
                    if "296" in ILX or "294" in ILX:
                        #print("Evento 296")
                        if "IRX" in ILX or "IRZ" in ILX:
                            print("Lleva la MFB-P2 DERECHA con terminación : 7216")
                            flag_mfbp2_der = True
                        for mod in csv.split(sep = ","):
                            if mod in pdcrVariantes["large"]:
                                flag_l = True
                            if mod in pdcrVariantes["medium"]:
                                flag_m = True
                            if mod in pdcrVariantes["small"]:
                                flag_s = True
                        print("\t\t+++++++++++ FLAGS de",ILX,":+++++++++++\n Flag S - ",flag_s," Flag M - ",flag_m," Flag L - ",flag_l," Flag MFB-P2 DER: ",flag_mfbp2_der)
                        if flag_mfbp2_der == True:
                            flag_mfbp2 = ["12975407216", True]
                        else:
                            flag_mfbp2 = ["12975407316", True]
                        if flag_l == True:
                            qr_pdcr = {
                                "PDC-R": ["12239061602", True],
                                "PDC-RMID": ["", False],
                                "PDC-RS": ["", False],
                                "PDC-D": ["12239060402", True],
                                "PDC-P": ["12239060702", True],
                                "MFB-P1": ["12975402001", True],
                                "MFB-S": ["12235403215", True],
                                "MFB-E": ["12975403015", True],
                                "MFB-P2": flag_mfbp2
                                }
                            print("Variante de caja PDC-R")
                        if flag_m == True and flag_l == False:
                            qr_pdcr = {
                                "PDC-R": ["", False],
                                "PDC-RMID": ["12239061502", True],
                                "PDC-RS": ["", False],
                                "PDC-D": ["12239060402", True],
                                "PDC-P": ["12239060702", True],
                                "MFB-P1": ["12975402001", True],
                                "MFB-S": ["12235403215", True],
                                "MFB-E": ["12975403015", True],
                                "MFB-P2": flag_mfbp2
                                }
                            print("Variante de caja PDC-RMID")
                        if flag_s == True and flag_m == False:
                            print("Variante de caja PDC-RS")
                            qr_pdcr = {
                                "PDC-R": ["", False],
                                "PDC-RMID": ["", False],
                                "PDC-RS": ["12239061402", True],
                                "PDC-D": ["12239060402", True],
                                "PDC-P": ["12239060702", True],
                                "MFB-P1": ["12975402001", True],
                                "MFB-S": ["12235403215", True],
                                "MFB-E": ["12975403015", True],
                                "MFB-P2": flag_mfbp2
                                }
                        if flag_s == False and flag_m == False and flag_l == False:
                            print("La caja no contiene módulos pertenecientes a las categorías.")
                            qr_pdcr = {
                                "PDC-R": ["", False],
                                "PDC-RMID": ["", False],
                                "PDC-RS": ["", False],
                                "PDC-D": ["12239060402", True],
                                "PDC-P": ["12239060702", True],
                                "MFB-P1": ["12975402001", True],
                                "MFB-S": ["12235403215", True],
                                "MFB-E": ["12975403015", True],
                                "MFB-P2": flag_mfbp2
                                }

                    temp = {
                        "DBEVENT": data,
                        "PEDIDO": ILX,
                        "DATETIME": "AUTO",
                        "MODULOS_VISION": {"INTERIOR": csv.split(sep = ",")},
                        "MODULOS_TORQUE": {"INTERIOR": csv.split(sep = ",")},
                        "MODULOS_ALTURA": {"INTERIOR": csv.split(sep = ",")},
                        "QR_BOXES": qr_pdcr,
                        "ACTIVE": 1
                        }
                    print("Códigos QR FINAL: ",qr_pdcr)
                    #print("ILX: ",ILX)
                    #print("Modulos que tiene: ",csv)
                    #print("Modulos que tiene TIPO: ",type(csv))
                    #print("Modulos que tiene el ILX: ",csv.split(","))
                    #print("Modulos que tiene convertido a array TIPO: ",type(csv.split(",")))
                    modulosDesconocidos = set(csv.split(",")) - set(modulesExisting["MODULO"])
                    modulosDesconocidos_t = set(csv.split(",")) - set(modulesExisting_t["MODULO"])
                    #print("Comparación; Modulos del ILX que NO están en la base de datos: ", modulosDesconocidos)
                    #print("Comparación; Modulos del ILX que NO están en la base de datos LEN VISION: ", len(modulosDesconocidos))
                    #print("Comparación; Modulos del ILX que NO están en la base de datos LEN TORQUES: ", len(modulosDesconocidos_t))
                    #print("Comparación tipo", type(modulosDesconocidos))
                    if len(modulosDesconocidos) == 0 and len(modulosDesconocidos_t) == 0:
                        modularities.append(temp)
                    else:
                        ilxfaltantes["ILX"][ILX] = {
                            "vision": [],
                            "torque": []
                            }
                        for e in modulosDesconocidos:
                            ilxfaltantes["ILX"][ILX]["vision"].append(e)
                        #print(e)
                            if not(e in modulosFaltantes):
                                modulosFaltantes.append(e)
                        for t in modulosDesconocidos_t:
                            #print(t)
                            ilxfaltantes["ILX"][ILX]["torque"].append(t)
                            if not(t in modulosFaltantes):
                                modulosFaltantes.append(t)
                        
                    ilxfaltantes["Modulos"] = modulosFaltantes
                    os.remove(root+'\\'+ file_name)
    #print("Lista total de Módulos Faltantes: ",ilxfaltantes)
    if len(modularities) != 0:
        updateModularities(modularities)
    return ilxfaltantes
 
def updateModularities(data):
    print("updating")
    #print("Data dentro de Upload Modularities: ",data)
    tabla = data[0]["DBEVENT"]
    print("TABLA en updating para DATS: ",tabla)
    endpoint = f"http://{host}:5000/api/get/{tabla}/pedidos/all/-/-/-/-/-"
    existing = requests.get(endpoint).json()
    if not("PEDIDO" in existing):
        existing["PEDIDO"] = []
    for i in data:
        try:
            if not(i["PEDIDO"] in existing["PEDIDO"]):
                endpoint = f"http://{host}:5000/api/post/pedidos"
                response = requests.post(endpoint, data = json.dumps(i))
            else:
                #pass
                index = existing["PEDIDO"].index(i["PEDIDO"])
                id = existing["ID"][index]
                endpoint = f"http://{host}:5000/api/update/pedidos/{id}"
                response = requests.post(endpoint, data = json.dumps(i))
        except Exception as ex:
            print (ex)

##################################### Determinantes management #################################
def refreshDeterminantes(data,usuario):
    data = makeDeterminantes(data,usuario)
    updateDeterminantes(data)

def makeDeterminantes(data,usuario):
    global determinantes
    determinantes = {
        "PDC-RS":[],
        "PDC-RMID":[],
        "PDC-R":[]
        }
    print("#################### Modulos Determinantes ####################")
    print("Modulos anteriormente cargados: ",determinantes)
    print("DATA que se pasa como arugmento a Determinantes",data)
    print("USUARIO que se pasa como arugmento a Determinantes",usuario)
    dir_path = os.path.join(os.getcwd(), '..\\determinantes\\')
    file_name = None
    for root, dirs, files in os.walk(dir_path):
        for file_name in files: 
            if file_name.endswith('.xls') or file_name.endswith('.xlsx'):
                file = openpyxl.load_workbook(filename = dir_path + file_name, data_only=True)
                sheets = file.sheetnames
                for sheet in sheets:
                    currentSheet = file[sheet]
                    columnas = ["PDC-RS","PDC-RMID","PDC-R"]
                    for variante in columnas:
                        if variante == "PDC-RS":
                            col = 2
                        if variante == "PDC-RMID":
                            col = 5
                        if variante == "PDC-R":
                            col = 8
                        #print("Col Actual: ",col)
                        for row in range(3, currentSheet.max_row + 1):
                            module = currentSheet.cell(column = col, row = row).value
                            if not(module in determinantes[variante]):
                                if module != None:
                                    determinantes[variante].append(module)
                                    #print("Modulo: ", module)
                print("Arreglo final de determinantes: ",determinantes)
                del file
                gc.collect()
                os.remove(root+'\\'+ file_name)

    structured_data = []
    for variante in determinantes:
        print("Variante: ",variante)
        for module in determinantes[variante]:
            print("Modulo: ",module)
            temp = {
            "DBEVENT": data,
            "MODULO": module,
            "VARIANTE": variante,
            "DATETIME": "AUTO",
            "USUARIO": usuario,
            "ACTIVE": 1
            }
            structured_data.append(temp)

    print ("\n total de modulos: ",len(structured_data))

    return structured_data

def updateDeterminantes(data):
    print("updating")
    tabla = data[0]["DBEVENT"]
    print("Update determinantes evento+-+-+-+-: ",tabla)
    endpoint = f"http://{host}:5000/api/get/{tabla}/definiciones/all/-/-/-/-/-"
    existing = requests.get(endpoint).json()
    if not("MODULO" in existing):
        existing["MODULO"] = []
    for i in data:
        try:
            if not(i["MODULO"] in existing["MODULO"]):
                endpoint = f"http://{host}:5000/api/post/definiciones"
                response = requests.post(endpoint, data = json.dumps(i))
            else:
                #pass
                index = existing["MODULO"].index(i["MODULO"])
                id = existing["ID"][index]
                endpoint = f"http://{host}:5000/api/update/definiciones/{id}"
                response = requests.post(endpoint, data = json.dumps(i))
        except Exception as ex:
            print (ex)

if __name__ == '__main__':
    #print("finished")
    refreshModules()
    makeModularities()
    #data = makeModules()
    #visionUpdate(data)
    #pdcrVariants("dumie")