from werkzeug.utils import secure_filename
from flask import Flask, request,  send_file, make_response
from openpyxl import Workbook
from openpyxl.chart.label import DataLabel, DataLabelList
from openpyxl.chart.series import SeriesLabel
from openpyxl.styles import Alignment, Font, PatternFill, NamedStyle,Fill
from openpyxl.chart import BarChart, Reference
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
import openpyxl
from datetime import datetime, timedelta, date, time
from flask_cors import CORS
from time import strftime
from pickle import load
import pymysql
import json
import os
import io
from os.path import exists  #para saber si existe una carpeta o archivo
from shutil import rmtree   #para eliminar carpeta con archivos dentro: rmtree("carpeta_con_archivos")
from os import remove       #para eliminar archivo único: remove("archivo.txt")
from os import rmdir        #para eliminar carpeta vacía: rmdir("carpeta_vacia")
import requests
from paho.mqtt import publish
import pyodbc
import auto_modularities
from auto_modularities import torques_value
from model import model

datos_conexion=model()
host,user,password,database,serverp2,dbp2,userp2,passwordp2,printerhost =datos_conexion.datos_acceso()

app = Flask(__name__)
CORS(app)
app.config['UPLOAD_FOLDER'] = os.path.join(os.getcwd(), '..\\')


@app.route("/server_famx/hora_servidor",methods=["GET"])
def servidorHora():
    try:
        connection = pyodbc.connect('DRIVER={SQL server}; SERVER='+serverp2+';DATABASE='+dbp2+';UID='+userp2+';PWD='+passwordp2)
        print("Conexión Éxitosa")
    except Exception as ex:
        print("Conexión a P2 Exception: ", ex)
        return {"exception": ex.args}
    query = "SELECT CURRENT_TIMESTAMP AS HORA_ACTUAL;"
    print("query: ",query)
    try:
        with connection.cursor() as cursor:
            items = cursor.execute(query)
            #result = cursor.fetchall()
            records = cursor.fetchall()
            insertObject = []
            columnNames = [column[0]
               for column in cursor.description
            ]
            for record in records:
               insertObject.append(dict(zip(columnNames, record)))
               #print("insertObject FINAL: ",insertObject)
            if len(insertObject) == 1:
                response = insertObject[0]
            elif len(insertObject) > 1:
                response = {}
                keys = list(insertObject[0])
                for key in keys:
                    response[key] = []
                    for item in insertObject:
                        response[key].append(item.pop(key))         
            else:
                response = {"items": 0}
            if "HORA_ACTUAL" in response:
                response["HORA_ACTUAL"] = response["HORA_ACTUAL"].strftime('%Y-%m-%d %H:%M:%S')
                print("response[HORA_ACTUAL]: ",response["HORA_ACTUAL"])
    except Exception as ex:
        print("myJsonResponse cursor Exception: ", ex)
        response = {"exception" : ex.args}
    finally:
        connection.close()
        return response
#####################################  Servicio para Etiquetas desde WEB ####################################
@app.route("/printer/etiqueta",methods=["POST"])
def etiqueta():
    response = {"items": 0}
    print("Dentro de Servicio para Etiqueta MANUAL")
    date = request.form['DATE']
    print("||||Fecha del registro: ",date)
    ref = request.form['REF']
    print("||||Referencia del registro: ",ref)
    hm = request.form['HM']
    print("||||HM del registro: ",hm)
    torques = request.form['TORQUES']
    print("||||Torques del registro: ",torques)
    #print("||||TIPO DE DATO: ",type(torques))
    torquesJson = json.loads(torques)
    #print("Torques convertido en JSON: ",torquesJson)
    #print("TIPO DE DATO: ",type(torquesJson))
    t_results_final = {}
    t_results = {}
    BoxIgnorar = ["MFB-S","MFB-E"] # Eliminar cuando se decida agregar el valor de estos torques a la estiquetas... además de modificarlo en el manager de visión para permitir dicha acción.
    for i in torquesJson:
        #print("i de Torques Json: ",i)
        #print("Valor de i: ",torquesJson[i])
        if i not in BoxIgnorar:
            t_results[i] = []
            for j in torquesJson[i]:
                #print("j dentro de torquesJson[i]: ",j)
                #print("valores internos del json: ",torquesJson[i][j])
                if torquesJson[i][j] == None:
                    torquesJson[i][j] = '-'
                t_results[i].append(torquesJson[i][j])
    #print("|||||||||Valores a agregar al valor FINAL: ",t_results)
    for i in t_results:
        #print("Caja: ",i," - Torques: ",t_results[i])
        if "PDC-R" in i:
            if t_results[i][0] != "-":
                #print("Este es el Bueno!")
                t_results_final["_PDC-R_"] = i+": "+str(t_results[i])
        else:
            t_results_final["_"+i+"_"] = i+": "+str(t_results[i])
    #print("|||||||||t_results FINAL: ",t_results_final)

    label = {
        "_DATE_":  date,
        "_REF_":   ref,
        "_QR_":    ref+" "+hm+" V.",
        "_TITLE_": " Vision-Torque-Altura Interior",
        "_HM_":    hm,
        "_RESULT_": "Fusibles y torques OK"
    }
    label.update(t_results_final)

    print("ETIQUETA:::::::::::::::::::::::::::::::::::::")
    print(label)
    #print("update(t_results_lbl): ",self.model.t_results_lbl)
    #print("_DATE_: ",self.model.datetime.strftime("%Y/%m/%d %H:%M:%S"))
    #print("_REF_: ",self.model.qr_codes["REF"])
    #print("_QR_: ",self.model.input_data["database"]["pedido"]["PEDIDO"])
    #print("_TITLE_: Vision-Torque-Altura Interior",)
    #print("_HM_: ",self.model.qr_codes["HM"])
    #print("_RESULT_: Fusibles y torques OK")
    try:
        #192.168.1.103 IP Maquina Vision
        publish.single("Printer/5", json.dumps(label), hostname= printerhost, qos = 2)
        response["items"] = 1
    except Exception as ex:
        print("ETIQUETA MANUAL Exception: ",ex)
        response = {"exception" : ex.args}
    finally:
        return response

#####################################  Upload Files Services ####################################
@app.route('/delete/filesmodularities', methods=['POST'])
def delRef():
    response = {"items": 0}
    try:
        path_carpeta = "..\\ILX";
        #se obtiene true si existe la carpeta
        existe_carpeta = os.path.isdir(path_carpeta)
        if existe_carpeta == True:
            try:
                #Eliminar la carpeta (con archivos dentro) anteriormente generada, (pueden quedarse por algún error de la matriz al tratar de cargar un formato inválido)
                #rmtree(path_carpeta)#para eliminar archivo único: from os import remove | remove("archivo.txt") ; para eliminar carpeta vacía: from os import rmdir | rmdir("carpeta_vacia")
                print("se elimina la carpeta")
                response = {"path" : 'Carpeta Eliminada desde la API,'}
            except OSError as error:
                print("ERROR AL ELIMINAR CARPETA:::\n",error)
                response = {"exception" : ex.args}
    except Exception as ex:
        print("uploadRef Exception: ", ex)
        response = {"exception" : ex.args}
        return response


@app.route('/upload/modularities', methods=['POST'])
def uploadRef():
    response = {"items": 0}
    allowed_file = False
    file = None
    try:
        if 'file' in request.files:
            file = request.files['file']
            if file.filename != '':
                filename = file.filename
                allowed_file = '.' in filename and \
                    filename.rsplit('.', 1)[1].lower() == "dat"
        if file and allowed_file:
            filename = secure_filename(file.filename)
            path = os.path.join(app.config['UPLOAD_FOLDER'], 'ILX')
            print(path, 'ACAAAAAAAA esta la ubicacion que se necesita subir')

            isExist = os.path.exists(path)
            if not isExist:
                # Create a new directory because it does not exist 
                os.makedirs(path)
                print("The new directory is created!", path)
            file.save(os.path.join(app.config['UPLOAD_FOLDER'], "ILX", filename))
            response["items"] = 1
    except Exception as ex:
        print("uploadRef Exception: ", ex)
        response = {"exception" : ex.args}
    finally:
        return response

@app.route('/update/modularities', methods=['POST'])
def updateRef():
    data = request.form['DBEVENT']
    print("DB a la que se cargan los DAT: ",data)
    ilxfaltantes = auto_modularities.makeModularities(data)
    return ilxfaltantes

@app.route('/update/modules', methods=['POST'])
def updateModules():
    response = {"items": 0}
    allowed_file = False
    file = None
    try:
        path_carpeta = "..\\modules";
        #se obtiene true si existe la carpeta
        existe_carpeta = os.path.isdir(path_carpeta)
        if existe_carpeta == True:
            try:
                #Eliminar la carpeta (con archivos dentro) anteriormente generada, (pueden quedarse por algún error de la matriz al tratar de cargar un formato inválido)
                rmtree(path_carpeta)#para eliminar archivo único: from os import remove | remove("archivo.txt") ; para eliminar carpeta vacía: from os import rmdir | rmdir("carpeta_vacia")
                print("se elimina la carpeta")
            except OSError as error:
                print("ERROR AL ELIMINAR CARPETA:::\n",error)

        data = request.form['DBEVENT']
        print("DB a la que se carga la Info: ",data)
        usuario = request.form['USUARIO']
        print("Usuario que carga la info: ",usuario)
        if 'file' in request.files:
            file = request.files['file']
            if file.filename != '':
                filename = file.filename
                allowed_file = '.' in filename and \
                    filename.rsplit('.', 1)[1].lower() in ['xls', 'xlsx']
        if file and allowed_file:
            filename = secure_filename(file.filename)
            path = os.path.join(app.config['UPLOAD_FOLDER'], "modules")
            #print(path, 'ACAAAAAAAA esta la ubicacion que se necesita subir')
            isExist = os.path.exists(path)
            if not isExist:
                # Create a new directory because it does not exist 
                os.makedirs(path)
                print("The new directory is created!", path)

            file.save(os.path.join(app.config['UPLOAD_FOLDER'], "modules", filename))
            auto_modularities.refreshModules(data)
            excelnew = {
                'DBEVENT': data,
                'ARCHIVO': filename,
                'USUARIO': usuario,
                'DATETIME': 'AUTO'
                }
            #print("Información que se manda al POST DE EVENTOS HISTORIAL: ",excelnew)
            endpoint = f"http://{host}:5000/api/post/historial"
            responseHistorial = requests.post(endpoint, data = json.dumps(excelnew))
            response["items"] = 1
    except Exception as ex:
        print("updateModules Exception: ", ex)
        response = {"exception" : ex.args}
    finally:
        return response

@app.route('/update/determinantes', methods=['POST'])
def updateDeterminantes():
    response = {"items": 0}
    allowed_file = False
    file = None
    try:
        path_carpeta = "..\\determinantes";
        #se obtiene true si existe la carpeta
        existe_carpeta = os.path.isdir(path_carpeta)
        if existe_carpeta == True:
            try:
                #Eliminar la carpeta (con archivos dentro) anteriormente generada, (pueden quedarse por algún error de la matriz al tratar de cargar un formato inválido)
                rmtree(path_carpeta)#para eliminar archivo único: from os import remove | remove("archivo.txt") ; para eliminar carpeta vacía: from os import rmdir | rmdir("carpeta_vacia")
                print("se elimina la carpeta")
            except OSError as error:
                print("ERROR AL ELIMINAR CARPETA:::\n",error)

        data = request.form['DBEVENT']
        print("DB a la que se carga la Info: ",data)
        usuario = request.form['USUARIO']
        print("Usuario que carga la info: ",usuario)
        if 'file' in request.files:
            file = request.files['file']
            if file.filename != '':
                filename = file.filename
                allowed_file = '.' in filename and \
                    filename.rsplit('.', 1)[1].lower() in ['xls', 'xlsx']
        if file and allowed_file:
            filename = secure_filename(file.filename)
            path = os.path.join(app.config['UPLOAD_FOLDER'], "determinantes")
            print(path, 'ACAAAAAAAA esta la ubicacion que se necesita subir')
            isExist = os.path.exists(path)
            if not isExist:
                # Create a new directory because it does not exist 
                os.makedirs(path)
                print("The new directory is created!", path)

            file.save(os.path.join(app.config['UPLOAD_FOLDER'], "determinantes", filename))
            auto_modularities.refreshDeterminantes(data,usuario)
            response["items"] = 1
    except Exception as ex:
        print("updateDeterminantes Exception: ", ex)
        response = {"exception" : ex.args}
    finally:
        return response

#########################################  CRUD Services ########################################
@app.route("/api/get/<table>/<column_1>/<operation_1>/<value_1>/<column_2>/<operation_2>/<value_2>",methods=["GET"])
def generalGET(table, column_1, operation_1, value_1, column_2, operation_2, value_2):
    if column_1=='all':
        query='SELECT * FROM ' +table+';'
    else:
        if value_2=='_':
            query = "SELECT * FROM " + table + " WHERE " + column_1 + operation_1 + "'{}';".format(value_1)
        else:
            query = "SELECT * FROM " + table + " WHERE " + column_1 + operation_1 + "'{}'".format(value_1)
            query += " AND " + column_2 + operation_2 + "'{}';".format(value_2)
    try:
        connection = pymysql.connect(host = host, user = user, passwd = password, database = database, cursorclass=pymysql.cursors.DictCursor)
    except Exception as ex:
        print("myJsonResponse connection Exception: ", ex)
        return {"exception": ex.args}
    try:
        with connection.cursor() as cursor:
            items = cursor.execute(query)
            result = cursor.fetchall()
            if len(result) == 1:
                response = result[0]
            elif len(result) > 1:
                response = {}
                keys = list(result[0])
                for key in keys:
                    response[key] = []
                    for item in result:
                        response[key].append(item.pop(key))         
            else:
                response = {"items": items}
    except Exception as ex:
        print("myJsonResponse cursor Exception: ", ex)
        response = {"exception" : ex.args}
    finally:
        connection.close()
        return response

@app.route("/api/get_col/<table>/<column_1>/<operation_1>/<value_1>/<column_2>/<operation_2>/<value_2>/<value_colums>/<value_QR>/<value_caja>",methods=["GET"])
def GET_column(table, column_1, operation_1, value_1, column_2, operation_2, value_2,value_colums,value_QR,value_caja):
    if column_1=='all' or value_colums=='all':
        query='SELECT * FROM ' +table+';'
    else:
        if value_2=='_':
            if value_colums=='all':
                query = "SELECT * FROM " + table + " WHERE " + column_1 + operation_1 + "'{}';".format(value_1)
            else:
                query = "SELECT "+ value_colums +" FROM " + table + " WHERE " + column_1 + operation_1 + "'{}';".format(value_1)
        else:
            if value_colums=='all':
                query = "SELECT * FROM " + table + " WHERE " + column_1 + operation_1 + "'{}'".format(value_1)
                query += " AND " + column_2 + operation_2 + "'{}';".format(value_2)
            else:
                query = "SELECT "+ value_colums +" FROM " + table + " WHERE " + column_1 + operation_1 + "'{}'".format(value_1)
                query += " AND " + column_2 + operation_2 + "'{}';".format(value_2)
    try:
        connection = pymysql.connect(host = host, user = user, passwd = password, database = database, cursorclass=pymysql.cursors.DictCursor)
    except Exception as ex:
        print("GET connection Exception: ", ex)
        return {"exception": ex.args}
    try:
        qr_repetido=False
        response_hms={}
        with connection.cursor() as cursor:
            items = cursor.execute(query)
            result = cursor.fetchall()
            if len(result) > 0:
                response = {}
                keys = list(result[0])
                for key in keys:
                    response[key] = []
                    response_hms[key] = []
                    for item in result:
                        item_dict=json.loads(item[key])
                        item_qr_box=""
                        if value_caja in item_dict.keys():
                            item_qr_box=item_dict[value_caja]
                            

                        if value_QR == item_qr_box:
                            qr_repetido=True
                            hm=item_dict["HM"]
                            response_hms[key].append(hm)
                            
                        
                        #response[key].append(item_qr_box)
                        item.pop(key)
                        #response[key].append(item.pop(key))           
            else:
                response = {"items": items}
    except Exception as ex:
        print("GET cursor Exception: ", ex)
        response = {"exception" : ex.args}
    finally:
        connection.close()
        
        if qr_repetido:
            response_ok= response_hms
        else:
            #response_ok=response
            response_ok={}
            response_ok["SERIALES"] = []
            response_ok["SERIALES"].append("NONE")
        print("responseee_ok",response_ok)
        return response_ok

@app.route('/query/get/<query>',methods=['GET'])
def query(query):
    print(query);
    try:
        connection = pymysql.connect(host = host, user = user, passwd = password, database = database, cursorclass=pymysql.cursors.DictCursor)
    except Exception as ex:
        print("myJsonResponse connection Exception: ", ex)
        return {"exception": ex.args}
    try:
        with connection.cursor() as cursor:
            items = cursor.execute(query)
            result = cursor.fetchall()
            print("el result es mejor",result)
            if len(result) > 0:
                response = {}
                keys = list(result[0])
                for key in keys:
                    response[key] = []
                    for item in result:
                        response[key].append(item.pop(key))   
                response["columns"] = keys
            else:
                response = {"items": items}
    except Exception as ex:
        print("myJsonResponse cursor Exception: ", ex)
        response = {"exception" : ex.args}
    finally:
        connection.close()
        return response





@app.route("/api/post/<table>",methods=["POST"])
def generalPOST(table):
    def escape_name(s):
        name = '`{}`'.format(s.replace('`', '``'))
        return name
    data = request.get_json(force=True)
    #print("Data -*-*-*--*-*-*-**: ",data)
    try:
        if ("DBEVENT" in data):
            #print("True SI HAY DBEVENT")
            print("DBEVENT: ",data["DBEVENT"])
            connection = pymysql.connect(host = host, user = user, passwd = password, database = data["DBEVENT"])
            del data["DBEVENT"]
        else:
            #print ("False NO HAY DBEVENT, TODO FLUYE NORMAL")
            connection = pymysql.connect(host = host, user = user, passwd = password, database = database)
    except Exception as ex:
        print("generalPOST connection Exception: ", ex)
        return {"exception": ex.args}
    try:
        query = "INSERT INTO " + table
        keys = list(data)
        cols = ', '.join(map(escape_name, keys))
        placeholders = ', '.join(['%({})s'.format(key) for key in keys])
        query += ' ({}) VALUES ({})'.format(cols, placeholders)
        for key in data:
            try:
                if key == "DATETIME":
                    if data[key] == "AUTO":
                        data[key] = datetime.now().isoformat()
                if type(data[key]) == dict:
                    data[key] = json.dumps(data[key])
            except Exception as ex:
                print("keys inspection Exception: ", ex)
        with connection.cursor() as cursor:
            items = cursor.execute(query, data)
        connection.commit()
        response = {"items": items}
    except Exception as ex:
        print("generalPOST Exception: ", ex)
        response = {"exception": ex.args}
    finally:
        connection.close()
        return response

@app.route("/api/delete/<table>/<int:ID>",methods=["POST"])
def delete(table, ID):
    try:
        connection = pymysql.connect(host = host, user = user, passwd = password, database = database)
    except Exception as ex:
        print("delete connection Exception: ", ex)
        return {"exception": ex.args}
    try:
        with connection.cursor() as cursor:
            items = cursor.execute(f"DELETE FROM {table} WHERE ID={ID}")
        connection.commit()
        response = {"items": items}
    except Exception as ex:
        print("dele Exception: ", ex)
        response = {"exception": ex.args}
    finally:
        connection.close()
        return response

@app.route("/api/update/<table>/<int:ID>",methods=["POST"])
def update(table, ID):
    def escape_name(s):
        name = '`{}`'.format(s.replace('`', '``'))
        return name
    data = request.get_json(force=True)
    try:
        if ("DBEVENT" in data):
            #print("True SI HAY DBEVENT")
            print("DBEVENT: ",data["DBEVENT"])
            connection = pymysql.connect(host = host, user = user, passwd = password, database = data["DBEVENT"])
            del data["DBEVENT"]
        else:
            #print ("False NO HAY DBEVENT, TODO FLUYE NORMAL")
            connection = pymysql.connect(host = host, user = user, passwd = password, database = database)
        #connection = pymysql.connect(host = host, user = user, passwd = password, database = database)
    except Exception as ex:
        print("generalPOST connection Exception: ", ex)
        return {"exception": ex.args}
    try:
        query = "UPDATE " + table + f" SET"
        for i in data:
            if i == "DATETIME":
                if data[i] == "AUTO":
                    data[i] = datetime.now().isoformat()
            key = escape_name(i)
            if type(data[i]) == dict:
                data[i] = json.dumps(data[i])
            query += f' {key}=%({i})s,'
        query = query[:-1]
        query += f" WHERE ID={ID}"
        with connection.cursor() as cursor:
            items = cursor.execute(query,data)
        connection.commit()
        response = {"items": items}
    except Exception as ex:
        print("delete Exception: ", ex)
        response = {"exception": ex.args}
    finally:
        connection.close()
        return response

@app.route("/api/get/pdcr/variantes",methods=["GET"])
def variantes():
    pdcrVariantes = {
    "small": [],
    "medium": [],
    "large": [],
    }
    endpoint = f"http://{host}:5000/api/get/definiciones/ACTIVE/=/1/_/_/_"
    pdcrVariantesDB = requests.get(endpoint).json()
    #print("pdcrVariantesDB-------",pdcrVariantesDB)
    if len(pdcrVariantesDB["MODULO"]) > 0:
        #print("Cantidad de Módulos: ",len(pdcrVariantesDB["MODULO"]))
        #print("Lista de Módulos: ",pdcrVariantesDB["MODULO"])
        #print("Lista de Variantes: ",pdcrVariantesDB["VARIANTE"])
        for i in pdcrVariantesDB["MODULO"]:
            #print("Modulo Actual (i)",i)
            #print("Index de Modulo Actual (i)",pdcrVariantesDB["MODULO"].index(i))
            #print("Variante correspondiente a Modulo Actual: ",pdcrVariantesDB["VARIANTE"][pdcrVariantesDB["MODULO"].index(i)])
            if pdcrVariantesDB["VARIANTE"][pdcrVariantesDB["MODULO"].index(i)] == "PDC-R":
                pdcrVariantes["large"].append(i)
                #print("ES UNA PDC-R LARGE")
            elif pdcrVariantesDB["VARIANTE"][pdcrVariantesDB["MODULO"].index(i)] == "PDC-RMID":
                #print("ES UNA PDC-R MEDIUM")
                pdcrVariantes["medium"].append(i)
            elif pdcrVariantesDB["VARIANTE"][pdcrVariantesDB["MODULO"].index(i)] == "PDC-RS":
                #print("ES UNA PDC-R SMALL")
                pdcrVariantes["small"].append(i)
    return pdcrVariantes

@app.route("/api/get/preview/modularity/<ILX>",methods=["GET"])
def preview(ILX):
    endpoint = f"http://{host}:5000/api/get/pdcr/variantes"
    pdcrVariantes = requests.get(endpoint).json()
    print("Lista Final de Variantes PDC-R: \n",pdcrVariantes)
    flag_l = False
    flag_m = False
    flag_s = False
    endpoint = f"http://{host}:5000/api/get/pedidos/PEDIDO/=/{ILX}/ACTIVE/=/1"
    response = requests.get(endpoint).json()
    #print("RESPONSE ",response)
    #print("RESPONSE ",response["MODULOS_VISION"])
    response_json = json.loads(response["MODULOS_VISION"])
    #print("RESPONSE JSON ",response_json)
    #print("RESPONSE JSON ",response_json["INTERIOR"][0])
    #arrayModules = response["MODULOS_FUSIBLES"][0].split(",")
    modules = response_json["INTERIOR"]
    print(f"\n\t\tMODULOS_FUSIBLES:\n{modules}")
    #print("Modulos SPLIT: ",arrayModules)
    modularity = {
        'vision': {
            'PDC-P': {},
            'PDC-D': {},
            'PDC-R': {},
            'PDC-RMID': {},
            'PDC-RS': {},
            'PDC-S': {}, 
            'PDC-S17': {}, 
            'PDC-S20': {}, 
            'PDC-S21': {}, 
            'PDC-S9': {}, 
            'PDC-S19': {}, 
            'TBLU': {}
        },
        'torque': {
            'PDC-P':{},
            'PDC-D':{},
            'MFB-P1':{},
            'MFB-S':{},
            'MFB-E':{},
            'MFB-P2':{},
            'PDC-R':{},
            'PDC-RMID':{},
            'PDC-RS': {},
            'BATTERY':{},
            'BATTERY-2':{},
            'BATTERY-3':{}
            },
        'variante': {}
    }
    cajas = modularity["vision"].keys()
    cajas_torque = modularity["torque"].keys()
    #print("CAJAS: ",cajas)
    for module in modules:
        if module in pdcrVariantes["large"]:
            flag_l = True
        if module in pdcrVariantes["medium"]:
            flag_m = True
        if module in pdcrVariantes["small"]:
            flag_s = True
        #print("Module i de la Lista: "+module)
        endpoint_Module= f"http://{host}:5000/api/get/modulos_fusibles/MODULO/=/{module}/_/=/_"
        #print("Endpoint del módulo"+endpoint_Module)
        resultado = requests.get(endpoint_Module).json()
        #print("Modulo Informacion",resultado)
        if "MODULO" in resultado:
            modulos_cant = resultado["MODULO"].split(sep = ",")
            #print("creacion de array: ",modulos_cant)
            if len(modulos_cant) == 1: 
                for j in resultado:
                    if j == "ID" or j == "MODULO" or j == "CAJA_6" or j == "CAJA_7" or j == "CAJA_8":
                        pass
                        #resultado[j] = resultado[j][0]
                    else:
                        #print(j)
                        resultado_json = json.loads(resultado[j])
                        #print(resultado_json)
                        #print(type(resultado_json))
                        for box in cajas:
                            #print("BOX: ",box)
                            if box in resultado_json:
                                #print("Si existe la caja dentro del JSON: ",resultado_json[box])
                                for k in resultado_json[box]:
                                    if resultado_json[box][k] != "vacio":
                                         #print("K: ",k)
                                         #print("Valor de la cavidad: ",resultado_json[box][k])
                                         modularity["vision"][box][k] = [resultado_json[box][k],module] 
                            else:
                                pass
    print("\t\t+++++++++++ FLAGS de",ILX,":+++++++++++\n Flag S - ",flag_s," Flag M - ",flag_m," Flag L - ",flag_l)
    if flag_l == True:
        variante = "PDC-R"
    if flag_m == True and flag_l == False:
        variante = "PDC-RMID"
    if flag_s == True and flag_m == False:
        variante = "PDC-RS"
    if flag_s == False and flag_m == False and flag_l == False:
        variante = "N/A"
        print("La caja no contiene módulos pertenecientes a las categorías.")
    modularity["variante"] = variante
    print("Variante de Caja: ",variante)
    #print("Response Modulos Torque: ",response["MODULOS_TORQUE"])
    response_torque = json.loads(response["MODULOS_TORQUE"])
    #print("Response Modulos Torque: ",response_torque)
    modules_torque = response_torque["INTERIOR"]
    #print("Response Modulos Torque: ",modules_torque)
    print(f"\n\t\tMODULOS_TORQUE:\n{modules_torque}")
    for modulet in modules_torque:
        #print("Module i de la Lista: "+module)
        endpoint_Modulet= f"http://{host}:5000/api/get/modulos_torques/MODULO/=/{modulet}/_/=/_"
        #print("Endpoint del módulo"+endpoint_Module)
        resultadot = requests.get(endpoint_Modulet).json()
        #print("Modulo Informacion",resultadot)
        if "MODULO" in resultadot:
            modulos_cant_t = resultadot["MODULO"].split(sep = ",")
            #print("creacion de array: ",modulos_cant_t)
            if len(modulos_cant_t) == 1: 
                for j in resultadot:
                    if j == "ID" or j == "MODULO":
                        pass
                        #resultado[j] = resultado[j][0]
                    else:
                        #print(j)
                        resultadot_json = json.loads(resultadot[j])
                        #print(resultado_json)
                        #print(type(resultado_json))
                        for box_torque in cajas_torque:
                            #print("BOX: ",box_torque)
                            if box_torque in resultadot_json:
                                #print("Si existe la caja dentro del JSON de Torques: ",resultadot_json[box_torque])
                                for k in resultadot_json[box_torque]:
                                    if resultadot_json[box_torque][k] == 1 or resultadot_json[box_torque][k] == True:
                                         #print("K: ",k)
                                         #print("Aplica torque?: ",resultadot_json[box_torque][k])
                                         modularity["torque"][box_torque][k] = [resultadot_json[box_torque][k],modulet]
                            else:
                                pass
    return modularity

################################################## Respaldos de Base de Datos Endpoint  ####################################################
@app.route("/api/get/bkup",methods=["GET"])
def bkup():
    items = {
        "status": False,
        "dir": "",
        "nombre": ""
        }
    ####### Cambiar Dirección de la carpeta destino donde se guardarán los Backups, dependiendo de la máquina o computadora en la que se correrá la API #######
    dest_folder = "C:/Users/administrador/Documents/MBI-EVTA-BKUPS/DATABASE"
    print("Petición de BACKUP")
    try:
        if os.path.isdir(dest_folder):
            print("La Carpeta para respaldos SI existe!")
            path = os.getcwd()   # show current working directory (cwd)
            print("path",path)
            os.chdir('C:/xampp/mysql/bin')
            filestamp = strftime('%Y%m%d-%H%M%S')
            filename = "%s/%s-%s.sql" % (dest_folder, filestamp, database)
            db_dump = "mysqldump --single-transaction -h " + host + " -u " + user + " -p" + password + " " + database + " > " + filename
            os.system(db_dump)
            items["status"] = True
            items["dir"] = filename
            items["nombre"] = filestamp+"-"+database
            print("DATABASE BACKUP EXITOSO")
        else:
            print("La Carpeta para respaldos NO existe!")
            items["dir"] = dest_folder
    except Exception as ex:
        print("DB BKUP Exception: ",ex)
    return items

################################################## Crear Base de Datos (Evento)  ####################################################
@app.route("/api/post/newEvent",methods=["POST"])
def newEvent():
    charSet = "utf8mb4_bin"
    historial = {
        "DBEVENT": "",
        "ARCHIVO": "",
        "USUARIO": "",
        "DATETIME": "",
    }
    activo = {
        "DBEVENT": "",
        "ACTIVE": ""
    }

    data = request.get_json(force=True)
    print("Data: ",data)
    event_name = 'evento_'+data["EVENTO"]+"_"+data["NUMERO"]+"_"+data["CONDUCCION"]
    historial["USUARIO"] = data["USUARIO"]
    historial["DATETIME"] = data["DATETIME"]
    historial["DBEVENT"] = event_name

    activo["ACTIVE"] = data["ACTIVE"]
    activo["DBEVENT"] = event_name
    try:
        connection = pymysql.connect(host = host, user = user, passwd = password)
    except Exception as ex:
        print("generalPOST connection Exception: ", ex)
        return {"exception": ex.args}
    try:
        with connection.cursor() as cursor:
            items = cursor.execute("create database "+event_name)
            sql = "use "+event_name
            cursor.execute(sql)
            definicionesTable = """CREATE TABLE definiciones (
            ID int primary key AUTO_INCREMENT, 
            MODULO text CHARACTER SET utf8mb4 COLLATE utf8mb4_bin NOT NULL, 
            VARIANTE text CHARACTER SET utf8mb4 COLLATE utf8mb4_bin NOT NULL,
            DATETIME datetime NOT NULL,
            USUARIO text CHARACTER SET utf8mb4 COLLATE utf8mb4_bin NOT NULL,
            ACTIVE tinyint NOT NULL
            )"""
            cursor.execute(definicionesTable)
            fusiblesTable = """CREATE TABLE modulos_fusibles (
            ID int primary key AUTO_INCREMENT, 
            MODULO text CHARACTER SET utf8mb4 COLLATE utf8mb4_bin NOT NULL, 
            CAJA_1 longtext CHARACTER SET utf8mb4 COLLATE utf8mb4_bin NOT NULL,
            CAJA_2 longtext CHARACTER SET utf8mb4 COLLATE utf8mb4_bin NOT NULL,
            CAJA_3 longtext CHARACTER SET utf8mb4 COLLATE utf8mb4_bin NOT NULL,
            CAJA_4 longtext CHARACTER SET utf8mb4 COLLATE utf8mb4_bin NOT NULL,
            CAJA_5 longtext CHARACTER SET utf8mb4 COLLATE utf8mb4_bin NOT NULL,
            CAJA_6 longtext CHARACTER SET utf8mb4 COLLATE utf8mb4_bin NOT NULL,
            CAJA_7 longtext CHARACTER SET utf8mb4 COLLATE utf8mb4_bin NOT NULL,
            CAJA_8 longtext CHARACTER SET utf8mb4 COLLATE utf8mb4_bin NOT NULL,
            CAJA_9 longtext CHARACTER SET utf8mb4 COLLATE utf8mb4_bin NOT NULL,
            CAJA_10 longtext CHARACTER SET utf8mb4 COLLATE utf8mb4_bin NOT NULL,
            CAJA_11 longtext CHARACTER SET utf8mb4 COLLATE utf8mb4_bin NOT NULL,
            CAJA_12 longtext CHARACTER SET utf8mb4 COLLATE utf8mb4_bin NOT NULL,
            CAJA_13 longtext CHARACTER SET utf8mb4 COLLATE utf8mb4_bin NOT NULL,
            CAJA_14 longtext CHARACTER SET utf8mb4 COLLATE utf8mb4_bin NOT NULL,
            CAJA_15 longtext CHARACTER SET utf8mb4 COLLATE utf8mb4_bin NOT NULL,
            CAJA_16 longtext CHARACTER SET utf8mb4 COLLATE utf8mb4_bin NOT NULL
            )"""
            cursor.execute(fusiblesTable)
            alturaTable = """CREATE TABLE modulos_alturas (
            ID int primary key AUTO_INCREMENT, 
            MODULO text CHARACTER SET utf8mb4 COLLATE utf8mb4_bin NOT NULL, 
            CAJA_1 longtext CHARACTER SET utf8mb4 COLLATE utf8mb4_bin NOT NULL,
            CAJA_2 longtext CHARACTER SET utf8mb4 COLLATE utf8mb4_bin NOT NULL,
            CAJA_3 longtext CHARACTER SET utf8mb4 COLLATE utf8mb4_bin NOT NULL,
            CAJA_4 longtext CHARACTER SET utf8mb4 COLLATE utf8mb4_bin NOT NULL,
            CAJA_5 longtext CHARACTER SET utf8mb4 COLLATE utf8mb4_bin NOT NULL,
            CAJA_6 longtext CHARACTER SET utf8mb4 COLLATE utf8mb4_bin NOT NULL,
            CAJA_7 longtext CHARACTER SET utf8mb4 COLLATE utf8mb4_bin NOT NULL,
            CAJA_8 longtext CHARACTER SET utf8mb4 COLLATE utf8mb4_bin NOT NULL,
            CAJA_9 longtext CHARACTER SET utf8mb4 COLLATE utf8mb4_bin NOT NULL,
            CAJA_10 longtext CHARACTER SET utf8mb4 COLLATE utf8mb4_bin NOT NULL,
            CAJA_11 longtext CHARACTER SET utf8mb4 COLLATE utf8mb4_bin NOT NULL,
            CAJA_12 longtext CHARACTER SET utf8mb4 COLLATE utf8mb4_bin NOT NULL,
            CAJA_13 longtext CHARACTER SET utf8mb4 COLLATE utf8mb4_bin NOT NULL,
            CAJA_14 longtext CHARACTER SET utf8mb4 COLLATE utf8mb4_bin NOT NULL,
            CAJA_15 longtext CHARACTER SET utf8mb4 COLLATE utf8mb4_bin NOT NULL,
            CAJA_16 longtext CHARACTER SET utf8mb4 COLLATE utf8mb4_bin NOT NULL
            )"""
            cursor.execute(alturaTable)
            torquesTable = """CREATE TABLE modulos_torques (
            ID int primary key AUTO_INCREMENT, 
            MODULO text CHARACTER SET utf8mb4 COLLATE utf8mb4_bin NOT NULL, 
            CAJA_1 longtext CHARACTER SET utf8mb4 COLLATE utf8mb4_bin NOT NULL,
            CAJA_2 longtext CHARACTER SET utf8mb4 COLLATE utf8mb4_bin NOT NULL,
            CAJA_3 longtext CHARACTER SET utf8mb4 COLLATE utf8mb4_bin NOT NULL,
            CAJA_4 longtext CHARACTER SET utf8mb4 COLLATE utf8mb4_bin NOT NULL,
            CAJA_5 longtext CHARACTER SET utf8mb4 COLLATE utf8mb4_bin NOT NULL,
            CAJA_6 longtext CHARACTER SET utf8mb4 COLLATE utf8mb4_bin NOT NULL,
            CAJA_7 longtext CHARACTER SET utf8mb4 COLLATE utf8mb4_bin NOT NULL,
            CAJA_8 longtext CHARACTER SET utf8mb4 COLLATE utf8mb4_bin NOT NULL,
            CAJA_9 longtext CHARACTER SET utf8mb4 COLLATE utf8mb4_bin NOT NULL,
            CAJA_10 longtext CHARACTER SET utf8mb4 COLLATE utf8mb4_bin NOT NULL,
            CAJA_11 longtext CHARACTER SET utf8mb4 COLLATE utf8mb4_bin NOT NULL,
            CAJA_12 longtext CHARACTER SET utf8mb4 COLLATE utf8mb4_bin NOT NULL,
            CAJA_13 longtext CHARACTER SET utf8mb4 COLLATE utf8mb4_bin NOT NULL,
            CAJA_14 longtext CHARACTER SET utf8mb4 COLLATE utf8mb4_bin NOT NULL,
            CAJA_15 longtext CHARACTER SET utf8mb4 COLLATE utf8mb4_bin NOT NULL,
            CAJA_16 longtext CHARACTER SET utf8mb4 COLLATE utf8mb4_bin NOT NULL
            )"""
            cursor.execute(torquesTable)
            pedidosTable = """CREATE TABLE pedidos (
            ID int primary key AUTO_INCREMENT, 
            PEDIDO text CHARACTER SET utf8mb4 COLLATE utf8mb4_bin NOT NULL,
            DATETIME datetime NOT NULL,
            MODULOS_VISION longtext CHARACTER SET utf8mb4 COLLATE utf8mb4_bin NOT NULL,
            MODULOS_TORQUE longtext CHARACTER SET utf8mb4 COLLATE utf8mb4_bin NOT NULL,
            MODULOS_ALTURA longtext CHARACTER SET utf8mb4 COLLATE utf8mb4_bin NOT NULL,
            QR_BOXES longtext CHARACTER SET utf8mb4 COLLATE utf8mb4_bin NOT NULL,
            ACTIVE tinyint NOT NULL
            )"""
            cursor.execute(pedidosTable)
            historialTable = """CREATE TABLE historial (
            ID int primary key AUTO_INCREMENT, 
            ARCHIVO longtext CHARACTER SET utf8mb4 COLLATE utf8mb4_bin NOT NULL, 
            USUARIO text CHARACTER SET utf8mb4 COLLATE utf8mb4_bin NOT NULL,
            DATETIME datetime NOT NULL
            )"""
            cursor.execute(historialTable)
            activoTable = """CREATE TABLE activo (
            ID int primary key AUTO_INCREMENT, 
            ACTIVE tinyint NOT NULL
            )"""
            cursor.execute(activoTable)
        connection.commit()
        response = {"items": items}
    except Exception as ex:
        print("generalPOST Exception: ", ex)
        response = {"exception": ex.args}
    finally:
        #print("Información que se manda al POST DE EVENTOS HISTORIAL: ",historial)
        endpoint = f"http://{host}:5000/api/post/historial"
        responseHistorial = requests.post(endpoint, data = json.dumps(historial))
        #print("Información que se manda al POST DE EVENTOS ACTIVO: ",activo)
        endpoint = f"http://{host}:5000/api/post/activo"
        responseActivo = requests.post(endpoint, data = json.dumps(activo))
        connection.close()
        return response

################################################## Eliminar Base de Datos (Evento)  ####################################################
@app.route("/api/delete/event",methods=["POST"])
def delEvent():
    charSet = "utf8mb4_bin"
    response = {"delete": 0}

    data = request.get_json(force=True)
    print("Data: ",data)
    #EVENTDELETE = data["DBEVENT"]
    try:
        connection = pymysql.connect(host = host, user = user, passwd = password, database = data["DBEVENT"])
    except Exception as ex:
        print("Delete Event connection Exception: ", ex)
        return {"exception": ex.args}
    try:
        with connection.cursor() as cursor:
            items = cursor.execute("DROP DATABASE "+data["DBEVENT"])
        connection.commit()
        response["delete"] = 1
    except Exception as ex:
        print("Delete Event Exception: ", ex)
        response = {"exception": ex.args}
    finally:
        connection.close()
        return response
################################################## Consultar Bases de Datos (Eventos)  ####################################################
@app.route("/api/get/eventos",methods=["GET"])
def eventos():
    lista = {
        "eventos": {}
        }
    try:
        connection = pymysql.connect(host = host, user = user, passwd = password)
    except Exception as ex:
        print("GET EVENTOS connection Exception: ", ex)
        return {"exception": ex.args}
    try:
        with connection.cursor() as cursor:
            items = cursor.execute("SHOW DATABASES")
            l = cursor.fetchall()
            #print ("Lista de dbs: ",l)
            x = []
            for i in l:
                #print("imprimiendo I 0 ",i[0])
                if 'evento' in i[0]:
                    #print("Este contiene evento: ",i[0])
                    x.extend(i)
                    
                    endpoint = f"http://{host}:5000/api/get/{i[0]}/historial/all/-/-/-/-/-"
                    respHistorial = requests.get(endpoint).json()
                    endpoint = f"http://{host}:5000/api/get/{i[0]}/activo/all/-/-/-/-/-"
                    respActivo = requests.get(endpoint).json()
                    #print("Respuesta de Historial: ",respHistorial)
                    #print("Respuesta de Historial Archivo: ",respHistorial["ARCHIVO"])
                    #print("Respuesta de Activo: ",respActivo)
                    #print("Respuesta de Activo: ",respActivo["ACTIVE"])
                    if type(respHistorial["ARCHIVO"]) == list:
                        #print("Es una lista!")
                        lista["eventos"][i[0]] = [respHistorial["ARCHIVO"][-1],respActivo["ACTIVE"]]
                    else:
                        #print("No es una lista, es posible que sea solo un elemento o esté vacío")
                        lista["eventos"][i[0]] = [respHistorial["ARCHIVO"],respActivo["ACTIVE"]]
            #print("Lista de bases de datos: ",x)
            print("Lista de eventos final: ",lista)
        connection.commit()
    except Exception as ex:
        print("GET EVENTOS Exception: ", ex)
    finally:
        connection.close()
        return lista

@app.route("/api/get/<db>/<table>/<column_1>/<operation_1>/<value_1>/<column_2>/<operation_2>/<value_2>",methods=["GET"])
def eventGET(table, db, column_1, operation_1, value_1, column_2, operation_2, value_2):
    if column_1=='all':
        query='SELECT * FROM ' +table+';'
    else:
        if value_2=='_':
            query = "SELECT * FROM " + table + " WHERE " + column_1 + operation_1 + "'{}';".format(value_1)
        else:
            query = "SELECT * FROM " + table + " WHERE " + column_1 + operation_1 + "'{}'".format(value_1)
            query += " AND " + column_2 + operation_2 + "'{}';".format(value_2)
    try:
        connection = pymysql.connect(host = host, user = user, passwd = password, database = db, cursorclass=pymysql.cursors.DictCursor)
    except Exception as ex:
        print("myJsonResponse connection Exception: ", ex)
        return {"exception": ex.args}
    try:
        with connection.cursor() as cursor:
            items = cursor.execute(query)
            result = cursor.fetchall()
            if len(result) == 1:
                response = result[0]
            elif len(result) > 1:
                response = {}
                keys = list(result[0])
                for key in keys:
                    response[key] = []
                    for item in result:
                        response[key].append(item.pop(key))         
            else:
                response = {"items": items}
    except Exception as ex:
        print("myJsonResponse cursor Exception: ", ex)
        response = {"exception" : ex.args}
    finally:
        connection.close()
        return response

@app.route("/api/get/<db>/preview/modularity/<ILX>",methods=["GET"])
def previewEvent(ILX,db):
    
    endpoint = f"http://{host}:5000/api/get/{db}/pdcr/variantes"
    pdcrVariantes = requests.get(endpoint).json()
    print("Lista Final de Variantes PDC-R: \n",pdcrVariantes)
    flag_l = False
    flag_m = False
    flag_s = False
    endpoint = f"http://{host}:5000/api/get/{db}/pedidos/PEDIDO/=/{ILX}/ACTIVE/=/1"
    response = requests.get(endpoint).json()
    #print("RESPONSE ",response)
    #print("RESPONSE ",response["MODULOS_VISION"])
    response_json = json.loads(response["MODULOS_VISION"])
    #print("RESPONSE JSON ",response_json)
    #print("RESPONSE JSON ",response_json["INTERIOR"][0])
    #arrayModules = response["MODULOS_FUSIBLES"][0].split(",")
    modules = response_json["INTERIOR"]
    print(f"\n\t\tMODULOS_FUSIBLES:\n{modules}")
    #print("Modulos SPLIT: ",arrayModules)
    modularity = {
        'vision': {
            'PDC-P': {},
            'PDC-D': {},
            'PDC-R': {},
            'PDC-RMID': {},
            'PDC-RS': {},
            'PDC-S': {}, 
            'PDC-S17': {}, 
            'PDC-S20': {}, 
            'PDC-S21': {}, 
            'PDC-S9': {}, 
            'PDC-S19': {}, 
            'TBLU': {},
            'F96': {},
            'F96-1': {}
        },
        'torque': {
            'PDC-P':{},
            'PDC-D':{},
            'MFB-P1':{},
            'MFB-S':{},
            'MFB-S2':{},
            'MFB-E':{},
            'MFB-P2':{},
            'PDC-R':{},
            'PDC-RMID':{},
            'PDC-RS': {},
            'BATTERY':{},
            'BATTERY-2':{},
            'BATTERY-3':{}
            },
        'variante': {}
    }
    cajas = modularity["vision"].keys()
    cajas_torque = modularity["torque"].keys()
    #print("CAJAS: ",cajas)
    for module in modules:
        if module in pdcrVariantes["large"]:
            flag_l = True
        if module in pdcrVariantes["medium"]:
            flag_m = True
        if module in pdcrVariantes["small"]:
            flag_s = True
        #print("Module i de la Lista: "+module)
        endpoint_Module= f"http://{host}:5000/api/get/{db}/modulos_fusibles/MODULO/=/{module}/_/=/_"
        #print("Endpoint del módulo"+endpoint_Module)
        resultado = requests.get(endpoint_Module).json()
        #print("Modulo Informacion",resultado)
        if "MODULO" in resultado:
            modulos_cant = resultado["MODULO"].split(sep = ",")
            #print("creacion de array: ",modulos_cant)
            if len(modulos_cant) == 1: 
                for j in resultado:
                    if j == "ID" or j == "MODULO":
                        pass
                        #resultado[j] = resultado[j][0]
                    else:
                        #print(j)
                        resultado_json = json.loads(resultado[j])
                        #print(resultado_json)
                        #print(type(resultado_json))
                        for box in cajas:
                            #print("BOX: ",box)
                            if box in resultado_json:
                                #print("Si existe la caja dentro del JSON: ",resultado_json[box])
                                for k in resultado_json[box]:
                                    if resultado_json[box][k] != "vacio":
                                         #print("K: ",k)
                                         #print("Valor de la cavidad: ",resultado_json[box][k])
                                         modularity["vision"][box][k] = [resultado_json[box][k],module] 
                            else:
                                pass
    print("\t\t+++++++++++ FLAGS de",ILX,":+++++++++++\n Flag S - ",flag_s," Flag M - ",flag_m," Flag L - ",flag_l)
    if flag_l == True:
        variante = "PDC-R"
    if flag_m == True and flag_l == False:
        variante = "PDC-RMID"
    if flag_s == True and flag_m == False:
        variante = "PDC-RS"
    if flag_s == False and flag_m == False and flag_l == False:
        variante = "N/A"
        print("La caja no contiene módulos pertenecientes a las categorías.")
    modularity["variante"] = variante
    print("Variante de Caja: ",variante)
    #print("Response Modulos Torque: ",response["MODULOS_TORQUE"])
    response_torque = json.loads(response["MODULOS_TORQUE"])
    #print("Response Modulos Torque: ",response_torque)
    modules_torque = response_torque["INTERIOR"]
    #print("Response Modulos Torque: ",modules_torque)
    print(f"\n\t\tMODULOS_TORQUE:\n{modules_torque}")
    for modulet in modules_torque:
        #print("Module i de la Lista: "+module)
        endpoint_Modulet= f"http://{host}:5000/api/get/{db}/modulos_torques/MODULO/=/{modulet}/_/=/_"
        #print("Endpoint del módulo"+endpoint_Module)
        resultadot = requests.get(endpoint_Modulet).json()
        #print("Modulo Informacion",resultadot)
        if "MODULO" in resultadot:
            modulos_cant_t = resultadot["MODULO"].split(sep = ",")
            #print("creacion de array: ",modulos_cant_t)
            if len(modulos_cant_t) == 1: 
                for j in resultadot:
                    if j == "ID" or j == "MODULO":
                        pass
                        #resultado[j] = resultado[j][0]
                    else:
                        #print(j)
                        resultadot_json = json.loads(resultadot[j])
                        #print(resultado_json)
                        #print(type(resultado_json))
                        for box_torque in cajas_torque:
                            #print("BOX: ",box_torque)
                            if box_torque in resultadot_json:
                                #print("Si existe la caja dentro del JSON de Torques: ",resultadot_json[box_torque])
                                for k in resultadot_json[box_torque]:
                                    if resultadot_json[box_torque][k] == 1 or resultadot_json[box_torque][k] == True:
                                         #print("K: ",k)
                                         #print("Aplica torque?: ",resultadot_json[box_torque][k])
                                         modularity["torque"][box_torque][k] = [resultadot_json[box_torque][k],modulet]
                            else:
                                pass
    return modularity

@app.route("/api/get/<db>/pdcr/variantes",methods=["GET"])
def variantesEvent(db):
    pdcrVariantes = {
    "small": [],
    "medium": [],
    "large": [],
    }
    endpoint = f"http://{host}:5000/api/get/{db}/definiciones/ACTIVE/=/1/_/_/_"
    pdcrVariantesDB = requests.get(endpoint).json()
    #print("pdcrVariantesDB-------",pdcrVariantesDB)
    try:
        if len(pdcrVariantesDB["MODULO"]) > 0:
            #print("Cantidad de Módulos: ",len(pdcrVariantesDB["MODULO"]))
            #print("Lista de Módulos: ",pdcrVariantesDB["MODULO"])
            #print("Lista de Variantes: ",pdcrVariantesDB["VARIANTE"])
            for i in pdcrVariantesDB["MODULO"]:
                #print("Modulo Actual (i)",i)
                #print("Index de Modulo Actual (i)",pdcrVariantesDB["MODULO"].index(i))
                #print("Variante correspondiente a Modulo Actual: ",pdcrVariantesDB["VARIANTE"][pdcrVariantesDB["MODULO"].index(i)])
                if pdcrVariantesDB["VARIANTE"][pdcrVariantesDB["MODULO"].index(i)] == "PDC-R":
                    pdcrVariantes["large"].append(i)
                    #print("ES UNA PDC-R LARGE")
                elif pdcrVariantesDB["VARIANTE"][pdcrVariantesDB["MODULO"].index(i)] == "PDC-RMID":
                    #print("ES UNA PDC-R MEDIUM")
                    pdcrVariantes["medium"].append(i)
                elif pdcrVariantesDB["VARIANTE"][pdcrVariantesDB["MODULO"].index(i)] == "PDC-RS":
                    #print("ES UNA PDC-R SMALL")
                    pdcrVariantes["small"].append(i)
    except Exception as ex:
        print("Variantes Exception: ", ex)
        return {"exception": ex.args}
    return pdcrVariantes

@app.route("/api/delete/<db>/<table>/<int:ID>",methods=["POST"])
def deleteEvent(table, ID,db):
    try:
        connection = pymysql.connect(host = host, user = user, passwd = password, database = db)
    except Exception as ex:
        print("delete connection Exception: ", ex)
        return {"exception": ex.args}
    try:
        with connection.cursor() as cursor:
            items = cursor.execute(f"DELETE FROM {table} WHERE ID={ID}")
        connection.commit()
        response = {"items": items}
    except Exception as ex:
        print("dele Exception: ", ex)
        response = {"exception": ex.args}
    finally:
        connection.close()
        return response

@app.route('/database/<db>/<table>/<column_of_table_1>/<operation_1>/<val_1>/<column_of_table_2>/<operation_2>/<val_2>',methods=['GET'])
def value_of_a_tableEvent(table,column_of_table_1,operation_1,val_1,column_of_table_2,operation_2,val_2,db):
    if column_of_table_1=='all':
        query='SELECT * FROM ' +table+';'
    else:
        if val_2=='_':
            query='SELECT * FROM ' +table+' WHERE '+column_of_table_1+operation_1+'"'+val_1+'";'
        else:
            query='SELECT * FROM ' +table+' WHERE '+column_of_table_1+operation_1+'"'+val_1+'" AND '+column_of_table_2+operation_2 +'"'+val_2+'";'
    print(query)
    #conexion con base de datos
    conexion =  pymysql.connect(host = host, user = user, passwd = password, database = db)
    cursor = conexion.cursor()
    cursor.execute(query)
    result = cursor.fetchone()
    
    if result == None:
        resp='NO HAY INFORMACION'
        response=resp
    else:
        resp='SI HAY INFORMACION'
        query = 'SELECT COLUMN_NAME FROM Information_Schema.Columns WHERE TABLE_NAME = ' + '"' + table + '";'
        cursor.execute(query)
        name_columns=cursor.fetchall()
        print(type(result))
        print(len(result))
        print(result)
        print(type(name_columns))
        print(len(name_columns))
        print(name_columns)

        dic={}
        for i in range(len(result)):
            dic[name_columns[i][0]]=result[i]
        print(dic)
        response=dic
    return response
################################################## Update Fijikura Server  ####################################################
#Servicio para verificar evento de refencia
@app.route("/nivel/get/<table>/<column_1>/<operation_1>/<value_1>/<column_2>/<operation_2>/<value_2>",methods=["GET"])
def nivelGET(table, column_1, operation_1, value_1, column_2, operation_2, value_2):
    #para llamar microservicio desde manager:
    #endpoint = ("http://{}/nivel/get/[agrucomb_prod].[dbo].[JP_estructura_combinacion_1er_nivel]/codigo_com_1er_nivel/=/ILX29420231008400/_/_/_".format(self.model.server))
    #endpoint = ("http://{}/nivel/get/[agrucomb_prod].[dbo].[JP_estructura_combinacion_1er_nivel]/all/_/_/_/_/_".format(self.model.server))
    #response = requests.get(endpoint).json()


    #petitión a la base de datos dependiendo de las variables que se den de entrada para el servicio
    if column_1=='all':
        query='SELECT * FROM ' +table+';'
    else:
        if value_2=='_':
            query = "SELECT * FROM " + table + " WHERE " + column_1 + operation_1 + "'{}';".format(value_1)
        else:
            query = "SELECT * FROM " + table + " WHERE " + column_1 + operation_1 + "'{}'".format(value_1)
            query += " AND " + column_2 + operation_2 + "'{}';".format(value_2)
    try:
        connection = pyodbc.connect('DRIVER={SQL server}; SERVER='+serverp2+';DATABASE='+dbp2+';UID='+userp2+';PWD='+passwordp2)
        print("Conexión Éxitosa")
    except Exception as ex:
        print("Conexión a P2 - Exception: ", ex)
        return {"exception": ex.args}


    #una vez que se realizó la petición GET, se obtiene un apuntador a la información que se quiere
    #y a partir de ahí se crea un diccionario con las dimensiones de la información requerida
    try:
        with connection.cursor() as cursor:
            items = cursor.execute(query)

            records = cursor.fetchall()
            
            print("records: \ntype: ")
            print(type(records))
            print(records)
            print("records[0]: \ntype: ")
            print(type(records[0]))
            print("_____________\n_____________\n_____________")
            print(records[0][-1])
            print("_____________\n_____________\n_____________")

            response = {"evento": str(records[0][-1]) }

    except Exception as ex:
        print("Exception: ", ex)
        response = {"items": 0}

    #se cierra la conexión y se regresa la respuesta "response"
    finally:
        connection.close()
        return response


@app.route("/seghm/get/<table>/<column_1>/<operation_1>/<value_1>/<column_2>/<operation_2>/<value_2>",methods=["GET"])
def famx2GET(table, column_1, operation_1, value_1, column_2, operation_2, value_2):
    if column_1=='all':
        query='SELECT * FROM ' +table+';'
    else:
        if value_2=='_':
            query = "SELECT * FROM " + table + " WHERE " + column_1 + operation_1 + "'{}';".format(value_1)
        else:
            query = "SELECT * FROM " + table + " WHERE " + column_1 + operation_1 + "'{}'".format(value_1)
            query += " AND " + column_2 + operation_2 + "'{}';".format(value_2)
    try:
        connection = pyodbc.connect('DRIVER={SQL server}; SERVER='+serverp2+';DATABASE='+dbp2+';UID='+userp2+';PWD='+passwordp2)
        print("Conexión Éxitosa")
    except Exception as ex:
        print("Conexión a P2 Exception: ", ex)
        return {"exception": ex.args}

    try:
        with connection.cursor() as cursor:
            items = cursor.execute(query)
            #result = cursor.fetchall()

            records = cursor.fetchall()
            insertObject = []
            columnNames = [column[0]
               for column in cursor.description
            ]

            for record in records:
               insertObject.append(dict(zip(columnNames, record)))
               #print("insertObject FINAL: ",insertObject)
            if len(insertObject) == 1:
                response = insertObject[0]
            elif len(insertObject) > 1:
                response = {}
                keys = list(insertObject[0])
                for key in keys:
                    response[key] = []
                    for item in insertObject:
                        response[key].append(item.pop(key))         
            else:
                response = {"items": 0}
    except Exception as ex:
        print("myJsonResponse cursor Exception: ", ex)
        response = {"exception" : ex.args}
    finally:
        connection.close()
        return response

@app.route("/seghm/update/<table>/<int:ID>",methods=["POST"])
def famx2update(table, ID):
    def escape_name(s):
        name = '`{}`'.format(s.replace('`', '``'))
        return name
    data = request.get_json(force=True)
    flag_torque = False
    flag_vision = False
    try:
        connection = pyodbc.connect('DRIVER={SQL server}; SERVER='+serverp2+';DATABASE='+dbp2+';UID='+userp2+';PWD='+passwordp2)
        print("|||| SERVICIO UPDATE Conexión Éxitosa")
    except Exception as ex:
        print("generalPOST connection Exception: ", ex)
        return {"exception": ex.args}
    try:
        query = "UPDATE " + table + f" SET"
        valores = []
        for key in data:
            try:
                if key == "DATETIME":
                    if data[key] == "AUTO":
                        data[key] = datetime.now().isoformat()
                if type(data[key]) == dict:
                    data[key] = json.dumps(data[key])
                valores.append(data[key])
            except Exception as ex:
                print("keys inspection Exception: ", ex)
            query += f' {key}= ?,'
            #print("primer Query: ",query)
            #print("Valores Final: ",valores)
        query = query[:-1]
        #print("query: ",query)
        query += f" WHERE ID={ID}"
        #print("query con += : ",query)
        with connection.cursor() as cursor:
            #print("dentro de cursor")
            items = cursor.execute(query, valores)
        connection.commit()
        response = {"items": 1}
    except Exception as ex:
        print("update Exception: ", ex)
        response = {"exception": 0}
    finally:
        connection.close()
        return response

@app.route("/seghm/post/<table>",methods=["POST"])
def famx2POST(table):
    def escape_name(s):
        name = '{}'.format(s.replace('`', '``'))
        return name
    data = request.get_json(force=True)
    #print("Data -*-*-*--*-*-*-**: ",data)
    try:
        connection = pyodbc.connect('DRIVER={SQL server}; SERVER='+serverp2+';DATABASE='+dbp2+';UID='+userp2+';PWD='+passwordp2)
        print("|||| SERVICIO POST FAMX2 Conexión Éxitosa")
    except Exception as ex:
        print("famx2POST connection Exception: ", ex)
        return {"exception": ex.args}
    try:
        query = "INSERT INTO " + table
        keys = list(data)
        #print("keys: ",keys)
        cols = ', '.join(map(escape_name, keys))
        placeholders = ', '.join(['?' for key in keys])
        query += ' ({}) VALUES ({})'.format(cols, placeholders)
        print("|||Query para POST: ",query)
        #print("Data: ",data)
        valores = []
        for key in data:
            try:
                if key == "DATETIME":
                    if data[key] == "AUTO":
                        data[key] = datetime.now().isoformat()
                if type(data[key]) == dict:
                    data[key] = json.dumps(data[key])
                valores.append(data[key])
            except Exception as ex:
                print("keys inspection Exception: ", ex)
        with connection.cursor() as cursor:
            items = cursor.execute(query, valores)
        connection.commit()
        response = {"items": 1} #Si el POST se realiza con éxito, al final regresará como respuesta el valor 1 asociado a la key "items"
    except Exception as ex:
        print("famx2POST Insert Exception: ", ex)
        response = {"exception": ex.args}
    finally:
        connection.close()
        return response


################################################## Webpages endpoints #########################################################
@app.route('/database/<table>/<column_of_table_1>/<operation_1>/<val_1>/<column_of_table_2>/<operation_2>/<val_2>',methods=['GET'])
def value_of_a_table(table,column_of_table_1,operation_1,val_1,column_of_table_2,operation_2,val_2):
    if column_of_table_1=='all':
        query='SELECT * FROM ' +table+';'
    else:
        if val_2=='_':
            query='SELECT * FROM ' +table+' WHERE '+column_of_table_1+operation_1+'"'+val_1+'";'
        else:
            query='SELECT * FROM ' +table+' WHERE '+column_of_table_1+operation_1+'"'+val_1+'" AND '+column_of_table_2+operation_2 +'"'+val_2+'";'
    print(query)
    #conexion con base de datos
    conexion =  pymysql.connect(host = host, user = user, passwd = password, database = database)
    cursor = conexion.cursor()
    cursor.execute(query)
    result = cursor.fetchone()
    
    if result == None:
        resp='NO HAY INFORMACION'
        response=resp
    else:
        resp='SI HAY INFORMACION'
        query = 'SELECT COLUMN_NAME FROM Information_Schema.Columns WHERE TABLE_NAME = ' + '"' + table + '";'
        cursor.execute(query)
        name_columns=cursor.fetchall()
        print(type(result))
        print(len(result))
        print(result)
        print(type(name_columns))
        print(len(name_columns))
        print(name_columns)

        dic={}
        for i in range(len(result)):
            dic[name_columns[i][0]]=result[i]
        print(dic)
        response=dic
    return response

@app.route('/json2/<table>/<column_of_table>/<operation_1>/<val_1>/<operation_2>/<val_2>',methods=['GET'])
def json2Return(table,column_of_table,operation_1,val_1,operation_2,val_2):
    items = 0

    if table == "availability":
        dic = {
            "columns":["datetime", "shift", "min", "percent", "pz"],
            "stop": {
                "datetime":[strftime("%d%b%Y-%H%M%S")]*21,
                "shift": [1]*21,
                "min": [21]*21,
                "cont":[21]*21,
                "percent": [21]*21,
                "pz": [21]*21
                },

            "pause": {
                "datetime":[strftime("%d%b%Y-%H%M%S")]*21,
                "shift": [2]*21,
                "min": [22]*21,
                "cont":[22]*21,
                "percent": [22]*21,
                "pz": [22]*21
                },

            "running": {
                "datetime":[strftime("%d%b%Y-%H%M%S")]*21,
                "shift": [3]*21,
                "min": [23]*21,
                "cont":[23]*21,
                "percent": [23]*21,
                "pz": [23]*21
                }      
            }
        return dic

    else:
        if column_of_table=='all':
            query='SELECT * FROM ' +table+';'
        else:
            if val_2=='_':
                query='SELECT * FROM ' +table+' WHERE '+column_of_table+operation_1+'"'+val_1+'";'
            else:
                query='SELECT * FROM ' +table+' WHERE '+column_of_table+operation_1+'"'+val_1+'" AND '+column_of_table+operation_2 +'"'+val_2+'";'
        # print(query)

        try:
            connection = pymysql.connect(host = host, user = user, passwd = password, database = database, cursorclass=pymysql.cursors.DictCursor)
        except Exception as ex:
            return {"exception": ex.args}
        try:
            with connection.cursor() as cursor:
                items = cursor.execute(query)
                result = cursor.fetchall()
                if len(result) > 0:
                    response = {}
                    keys = list(result[0])
                    for key in keys:
                        response[key] = []
                        for item in result:
                            response[key].append(item.pop(key))   
                    response["columns"] = keys
                else:
                    response = {"items": items}
            
        except Exception as ex:
            return {"exception": ex.args}
        return response

@app.route('/database/<table>/<column_of_table_1>/<operation_1>/<val_1>/<column_of_table_2>/<operation_2>/<val_2>/multi',methods=['GET'])
def value_of_a_table_2(table,column_of_table_1,operation_1,val_1,column_of_table_2,operation_2,val_2):
    if column_of_table_1=='all':
        query='SELECT * FROM ' +table+';'
    else:
        if val_2=='_':
            query='SELECT * FROM ' +table+' WHERE '+column_of_table_1+operation_1+'"'+val_1+'";'
        else:
            query='SELECT * FROM ' +table+' WHERE '+column_of_table_1+operation_1+'"'+val_1+'" AND '+column_of_table_2+operation_2 +'"'+val_2+'";'
    print(query)
    #conexion con base de datos
    conexion =  pymysql.connect(host = host, user = user, passwd = password, database = database)
    cursor = conexion.cursor()
    cursor.execute(query)
    result = cursor.fetchall()
    print(result)
    if result == ():
        resp='NO HAY INFORMACION'
        response=resp
    else:
        resp='SI HAY INFORMACION'
        query = 'SELECT COLUMN_NAME FROM Information_Schema.Columns WHERE TABLE_NAME = ' + '"' + table + '";'
        cursor.execute(query)
        #name_columns=cursor.fetchall()
        name_columns=cursor.fetchmany(len(result[0]))
        #print(type(result))
        #print(len(result))
        #print(result)
        #print(type(name_columns))
        #print(len(name_columns))
        #print(name_columns)
        #print(len(result[0]))
        dic={}
        #a=[]
        #for j in range(len(result)):
            #a.append(result[j][0])
        #print(a)
        
        for i in range(len(name_columns)):
            dic[name_columns[i][0]]=[]
            for j in range(len(result)):
                dic[name_columns[i][0]].append(result[j][i])
        print(dic['ID'])
        response=dic
    return response

@app.route('/info/<arnes>/<type_pts>/<caja>',methods=['GET'])
def info_cajas(arnes,type_pts,caja):
    path = "data/points/"
    file_name= path + caja+"_puntos_"+type_pts
    print(file_name)
    if arnes=="interior":
        print("interior")
        #file_name='test'
        #if var_pdcr==1:
        #    file_name='puntos_vision_caja_'+str(caja)+'_1'
        #print(file_name)
        with open(file_name, "rb") as f:
            pts= load(f)
            print(pts)
            print(len(pts))
        dic={"puntos":pts}
    if arnes=="motor":
        print("motor")
        #file_name='test'
        #if var_pdcr==1:
        #    file_name='puntos_vision_caja_'+str(caja)+'_1'
        #print(file_name)
        with open(file_name, "rb") as f:
            pts= load(f)
            print(pts)
            print(len(pts))
        dic={"puntos":pts}
    return dic

########################################################################################################################################
########################################################################################################################################
@app.route('/contar/<table>/<column>', methods=['GET'])
def data_count(table, column):
    turnos = request.get_json(force=True)
    turnos = {
            "1":["07-00","16-59"],
            "2":["17-00","06-59"],
            }

    print("turnos:",turnos)
    ####### REVISAR EN QUÉ TURNO ESTÁ LA HORA ACTUAL
    for elemento in turnos:

        hora_iniciostr = turnos[elemento][0]
        hora_finstr = turnos[elemento][1]

        inicio_split = hora_iniciostr.split("-")
        hora_inicio = int(inicio_split[0])
        minuto_inicio = int(inicio_split[1])
            
        fin_split = hora_finstr.split("-")
        hora_fin = int(fin_split[0])
        minuto_fin = int(fin_split[1])

            
        #Se obtiene la Hora actual (int)
        horaActual = datetime.now().hour
        #Minutos Actuales
        minActual = datetime.now().minute
            

        #se detecta el tipo de jornada ....
        caso1 = False #inicio menor que fin (jornada normal)
        caso2 = False #fin menor que inicio (jornada con cambio de día)
            
        if hora_inicio < hora_fin:
            #print("hora inicio menor que hora fin")
            caso1 = True
        else:
            if hora_inicio == hora_fin:
                #print("hora inicio = a hora fin")
                if minuto_inicio < minuto_fin:
                    #print("minuto inicio menor que minuto fin")
                    caso1 = True
                else:
                    #print("minuto inicio mayor que minuto fin")
                    caso2 = True
            else:
                #print("hora inciio menor que hora fin")
                caso2 = True


        #Fecha Actual
        fechaActual = datetime.today()
        ##Segundos Actuales
        ##secActual = datetime.now().second
        #delta time de un día
        td = timedelta(days = 1)
        ayerfechaActual = fechaActual - td
        mañanafechaActual = fechaActual + td

        hoy_year =  datetime.now().year
        hoy_month = datetime.now().month
        hoy_day =   datetime.now().day

        ayer_year =  ayerfechaActual.year
        ayer_month = ayerfechaActual.month
        ayer_day =   ayerfechaActual.day

        mañana_year =  mañanafechaActual.year
        mañana_month = mañanafechaActual.month
        mañana_day =   mañanafechaActual.day

        inicio_query = ""
        fin_query = ""

        #AQUÍ YA SE SABE EL TIPO DE HORARIO QUE SE ESTÁ REVISANDO, 
        #HAY QUE VER SI LA HORA ACTUAL ESTÁ DENTRO DE ESTE HORARIO

        if caso1 == True:

            init_date = datetime(hoy_year, hoy_month, hoy_day, hora_inicio, minuto_inicio )
            end_date  = datetime(hoy_year, hoy_month, hoy_day, hora_fin,    minuto_fin )

            if init_date <= fechaActual <= end_date:
                inicio_query = str(init_date.strftime('%Y-%m-%d-%H-%M'))
                fin_query =     str(end_date.strftime('%Y-%m-%d-%H-%M'))
                break

        if caso2 == True:

            init_date1 = datetime(    hoy_year,     hoy_month,     hoy_day,  hora_inicio,  minuto_inicio )
            end_date1 =  datetime( mañana_year,  mañana_month,  mañana_day,  hora_fin,     minuto_fin )
                
            init_date2 = datetime(ayer_year, ayer_month, ayer_day, hora_inicio, minuto_inicio )
            end_date2 =  datetime( hoy_year,  hoy_month,  hoy_day,    hora_fin,    minuto_fin )

            if init_date1 <= fechaActual <= end_date1:
                inicio_query = str(init_date1.strftime('%Y-%m-%d-%H-%M'))
                fin_query =     str(end_date1.strftime('%Y-%m-%d-%H-%M'))
                break

            if init_date2 <= fechaActual <= end_date2:
                inicio_query = str(init_date2.strftime('%Y-%m-%d-%H-%M'))
                fin_query =     str(end_date2.strftime('%Y-%m-%d-%H-%M'))
                break

    ####### CONTAR LOS ARNESES QUE HAY o HA HABIDO ENTRE TAL FECHA Y TAL FECHA DEPENDIENDO DEL CASO

    print("--------------------inicio_query: ",inicio_query)
    print("--------------------   fin_query: ",fin_query)

    query= "SELECT * FROM " +table+" WHERE "+ column + ">=" + "'" + inicio_query + "' AND " + column + "<=" + "'" + fin_query + "';"
    print("query: ",query)

    try:
        connection = pymysql.connect(host = host, user = user, passwd = password, database = database, cursorclass=pymysql.cursors.DictCursor)
    
    except Exception as ex:
        print("data_count connection Exception: ", ex)
        response = {"conteo" : 0}
        return response


    try:
        cursor = connection.cursor()
        cursor.execute(query)
        result = cursor.fetchall() 
        
        pedidos = []

        for i in result: ##Buscando diferentes Valores en el rango de fecha
            indice = result.index(i)
            if result[indice]["RESULTADO"] > 0: ##Revisando si existen resets
                pedidos.append(result[indice]["RESULTADO"])
                #print(pedidos)
        mylist = list(dict.fromkeys(pedidos)) ## Eliminando valores duplicados
        #print(len(pedidos), 'The Big ONE')
        ###CONTAR LOS ITEMS LEÍDOS

        #if inicio_query == "":
        #    conteo = 0
        #else:
        #    if len(result):
        #        if isinstance(result, str):
        #            conteo = 1
        #        if isinstance(result, list):
        #            conteo = len(result)
        #            #print(result)
        #    else:
        #        conteo = 0

        response = {"conteo" : len(pedidos)}

    except Exception as ex:
        print("data_count cursor Exception: ", ex)
        response = {"conteo" :0}
        return response

    finally:
        connection.close()
        return response

@app.route('/horaxhora/<table>/<column>', methods=['GET'])
def horaxhora(table, column):
    turnos = request.get_json(force=True)
    turnos = {
            "1":["07-00","16-59"],
            "2":["17-00","06-59"],
            }

    print("turnos:",turnos)
    ####### REVISAR EN QUÉ TURNO ESTÁ LA HORA ACTUAL
    for elemento in turnos:

        hora_iniciostr = turnos[elemento][0]
        hora_finstr = turnos[elemento][1]

        inicio_split = hora_iniciostr.split("-")
        hora_inicio = int(inicio_split[0])
        minuto_inicio = int(inicio_split[1])
            
        fin_split = hora_finstr.split("-")
        hora_fin = int(fin_split[0])
        minuto_fin = int(fin_split[1])

            
        #Se obtiene la Hora actual (int)
        horaActual = datetime.now().hour
        #Minutos Actuales
        minActual = datetime.now().minute
            

        #se detecta el tipo de jornada ....
        caso1 = False #inicio menor que fin (jornada normal)
        caso2 = False #fin menor que inicio (jornada con cambio de día)
            
        if hora_inicio < hora_fin:
            #print("hora inicio menor que hora fin")
            caso1 = True
        else:
            if hora_inicio == hora_fin:
                #print("hora inicio = a hora fin")
                if minuto_inicio < minuto_fin:
                    #print("minuto inicio menor que minuto fin")
                    caso1 = True
                else:
                    #print("minuto inicio mayor que minuto fin")
                    caso2 = True
            else:
                #print("hora inciio menor que hora fin")
                caso2 = True


        #Fecha Actual
        fechaActual = datetime.today()
        ##Segundos Actuales
        ##secActual = datetime.now().second
        #delta time de un día
        td = timedelta(days = 1)
        ayerfechaActual = fechaActual - td
        mañanafechaActual = fechaActual + td

        hoy_year =  datetime.now().year
        hoy_month = datetime.now().month
        hoy_day =   datetime.now().day

        ayer_year =  ayerfechaActual.year
        ayer_month = ayerfechaActual.month
        ayer_day =   ayerfechaActual.day

        mañana_year =  mañanafechaActual.year
        mañana_month = mañanafechaActual.month
        mañana_day =   mañanafechaActual.day

        inicio_query = ""
        fin_query = ""

        #AQUÍ YA SE SABE EL TIPO DE HORARIO QUE SE ESTÁ REVISANDO, 
        #HAY QUE VER SI LA HORA ACTUAL ESTÁ DENTRO DE ESTE HORARIO

        if caso1 == True:

            init_date = datetime(hoy_year, hoy_month, hoy_day, hora_inicio, minuto_inicio )
            end_date  = datetime(hoy_year, hoy_month, hoy_day, hora_fin,    minuto_fin )

            if init_date <= fechaActual <= end_date:
                inicio_query = str(init_date.strftime('%Y-%m-%d-%H-%M'))
                fin_query =     str(end_date.strftime('%Y-%m-%d-%H-%M'))
                break

        if caso2 == True:

            init_date1 = datetime(    hoy_year,     hoy_month,     hoy_day,  hora_inicio,  minuto_inicio )
            end_date1 =  datetime( mañana_year,  mañana_month,  mañana_day,  hora_fin,     minuto_fin )
                
            init_date2 = datetime(ayer_year, ayer_month, ayer_day, hora_inicio, minuto_inicio )
            end_date2 =  datetime( hoy_year,  hoy_month,  hoy_day,    hora_fin,    minuto_fin )

            if init_date1 <= fechaActual <= end_date1:
                inicio_query = str(init_date1.strftime('%Y-%m-%d-%H-%M'))
                fin_query =     str(end_date1.strftime('%Y-%m-%d-%H-%M'))
                break

            if init_date2 <= fechaActual <= end_date2:
                inicio_query = str(init_date2.strftime('%Y-%m-%d-%H-%M'))
                fin_query =     str(end_date2.strftime('%Y-%m-%d-%H-%M'))
                break

    ####### CONTAR LOS ARNESES QUE HAY o HA HABIDO ENTRE TAL FECHA Y TAL FECHA DEPENDIENDO DEL CASO

    print("--------------------inicio_query: ",inicio_query)
    print("--------------------   fin_query: ",fin_query)

    query= "SELECT HM,RESULTADO,INICIO,FIN,USUARIO FROM " +table+" WHERE "+ column + ">=" + "'" + inicio_query + "' AND " + column + "<=" + "'" + fin_query + "';"
    print("query: ",query)

    try:
        connection = pymysql.connect(host = host, user = user, passwd = password, database = database, cursorclass=pymysql.cursors.DictCursor)
    
    except Exception as ex:
        print("data_count connection Exception: ", ex)
        response = {'HM': 0,
                   'INICIO': 0,
                   'FIN': 0,
                   'RESULTADO': 0}
        return response


    try:
        cursor = connection.cursor()
        cursor.execute(query)
        result = cursor.fetchall() 
        print("result",result)
        if len(result) > 0:
            response = {}
            keys = list(result[0])
            for key in keys:
                response[key] = []
                for item in result:
                    response[key].append(item.pop(key))   
            response["columns"] = keys
        else:
            response ={'HM': 0,
                   'INICIO': 0,
                   'FIN': 0,
                   'RESULTADO': 0}
        

    except Exception as ex:
        print("data_count cursor Exception: ", ex)
        response = {'HM': 0,
                   'INICIO': 0,
                   'FIN': 0,
                   'RESULTADO': 0}
        return response

    finally:
        connection.close()
        return response


########                                #######
########    Area de consulta de datos   #######
########                                #######
@app.route('/descargar/<db>/<table>/<task>')
def descargar(db, table, task):
    query = 'SELECT * FROM ' +table+' WHERE '+task+';'
    print(query)
    try:
        connection = pymysql.connect(host = host, user = user, passwd = password, database = database, cursorclass=pymysql.cursors.DictCursor)
    except Exception as ex:
        print("myJsonResponse connection Exception: ", ex)
        return {"exception": ex.args}
    try:
        with connection.cursor() as cursor:
            items = cursor.execute(query)
            result = cursor.fetchall()
            #print("result: ",result)
            #print(result[0].keys())
            arreglo = []
            hourEnd = []
            h =  result[0]
            valores_fila = ','.join( str(valor) for valor in h)
            #print(valores_fila)
            li = list(valores_fila.split(","))
            #print(li)
            li.remove('VISION')
            li.remove('TORQUE')
            li.remove('ALTURA')
            li.remove('INTENTOS_VA')
            li.remove('INTENTOS_T')
            li.remove('SCRAP')
            li.remove('SERIALES')
            #li.remove('NOTAS')
            li.remove('ANGULO')
            #Capitalizando Titulos,   TITULO -> Titulo
            capt = []
            for l in li: 
                c = l.capitalize()
                capt.append(c)
                
            arreglo_intt = [] #Arreglo intentos torque
            arreglo_ang = [] #Arreglo Angulo
            arreglo_tor = [] #Arreglo Torque
            arreglo.append(capt)
            #arreglo.append(valores_fila)
            # Extraer subencabezados en una lista
            subencabezados = [subdict.keys() for subdict in torques_value.values()]

            # Convertir la lista de diccionarios a una lista plana de subencabezados
            subencabezados = [item for sublist in subencabezados for item in sublist]

            subencabezados.insert(0,'HM')
            subencabezados.insert(0,'ID')
            arreglo_intt.append(subencabezados)
            arreglo_ang.append(subencabezados)
            arreglo_tor.append(subencabezados)
            # print(len(subencabezados))
            # print(subencabezados)


            if len(result) > 0:
                # Procesar los resultados por fila
                columnas = []
                for fila in result:
                    del fila['VISION']
                    #del fila['TORQUE']
                    del fila['ALTURA']
                    del fila['INTENTOS_VA']
                    #del fila['INTENTOS_T']
                    del fila['SCRAP']
                    del fila['SERIALES']
                    #del fila['NOTAS']
                    # del fila['ANGULO']

                    dato = []
                    dato_intt = []
                    dato_tor = []
                    dato_ang = []
                    cajas =[]
                    #print(fila)
                    if fila["INTENTOS_T"]:
                        fila_json = json.loads(fila["INTENTOS_T"])
                        dato_intt.append(fila["ID"])
                        dato_intt.append(fila["HM"])
                        #print(fila["INTENTOS_T"])
                        
                        for i in torques_value:

                            for j in torques_value[i]:
                                #print(i)
                                #print(j)
                                if i in fila_json:
                                    try:
                                        valor = fila_json[i][j]
                                        caja = i
                                        #print("El valor existe:", valor)
                                        cajas.append(caja)
                                        dato_intt.append(valor)
                                    except IndexError:
                                        print("El índice j está fuera de rango para fila['INTENTOS_T'][i].")
                                else: 
                                    valor = None
                                    caja = i 
                                    cajas.append(caja)
                                    dato_intt.append(valor)

                            #print(dato_intt)
                            #print(len(dato_intt))
                    if fila["ANGULO"]:
                        fila_json = json.loads(fila["ANGULO"])
                        dato_ang.append(fila["ID"])
                        dato_ang.append(fila["HM"])
                        #print(fila["ANGULO"])
                        
                        for i in torques_value:

                            for j in torques_value[i]:
                                #print(i)
                                #print(j)
                                if i in fila_json:
                                    try:
                                        valor = fila_json[i][j]
                                        caja = i
                                        #print("El valor existe:", valor)
                                        #cajas.append(caja)
                                        dato_ang.append(valor)

                                    except IndexError:
                                        print("El índice j está fuera de rango para fila['ANGULO'][i].")
                                else: 
                                    valor = None
                                    caja = i
                                    #cajas.append(caja)
                                    dato_ang.append(valor)
                            #print(dato_ang)
                            #print(len(dato_ang))
                    if fila["TORQUE"]:
                        fila_json = json.loads(fila["TORQUE"])
                        dato_tor.append(fila["ID"])
                        dato_tor.append(fila["HM"])
                        #print(fila["TORQUE"])
                        
                        for i in torques_value:

                            for j in torques_value[i]:
                                #print(i)
                                #print(j)
                                if i in fila_json:
                                    try:
                                        valor = fila_json[i][j]
                                        caja = i
                                        #print("El valor existe:", valor)
                                        #cajas.append(caja)
                                        dato_tor.append(valor)

                                    except IndexError:
                                        print("El índice j está fuera de rango para fila['TORQUE'][i].")
                                else: 
                                    valor = None
                                    caja = i
                                    #cajas.append(caja)
                                    dato_tor.append(valor)
                            #print(dato_tor)
                            #print(len(dato_tor))
                                        
                    if fila["HM"] != 'HM000000000003':
                        arreglo.append(dato)
                        #print(fila)
                        for i in fila:
                            if 'INTENTOS_T' in i or 'ANGULO' in i or 'TORQUE' in i:
                                False
                            elif 'NOTAS' in i:
                                Notepad = json.loads(fila[i])
                                #print(Notepad['TORQUE'][1]) 
                                if 'TORQUE' in Notepad:
                                    dato.append(Notepad['TORQUE'][1])

                                elif 'VISION' in Notepad:
                                    dato.append(Notepad['VISION'][1])

                                else:                                   
                                    dato.append(fila[i])
                            else: 
                                #print(fila["FIN"])
                                # Define las dos fechas como cadenas de texto
                                hourEnd.append(fila['FIN'])
                                # Convierte las cadenas de texto en objetos datetime
                                #fecha1 = datetime.strptime(fecha1_str, "%d-%m-%Y %H:%M:%S")
                                #fecha2 = datetime.strptime(fecha2_str, "%d-%m-%Y %H:%M:%S")

                                dato.append(fila[i])

                    #arreglo_intt.append(cajas)
                    arreglo_intt.append(dato_intt)
                    arreglo_ang.append(dato_ang)
                    arreglo_tor.append(dato_tor)
                    #print(arreglo_intt)

            else:
                response = {"items": items}
                #print(response)

    except Exception as ex:
        print("myJsonResponse cursor Exception: ", ex)
        response = {"exception" : ex.args}


    #######################  REALIZANDO FORMATO EXCEL ################
    #######################                           ################

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Historial'





    # Crea una nueva hoja en el libro
    sheet2 = workbook.create_sheet("IntentosT") # insertando tabla por numero de posicion de izquierda a derecha
    sheet2 = workbook["IntentosT"]

    sheet3 = workbook.create_sheet("Angulo") # insertando tabla por numero de posicion de izquierda a derecha
    sheet3 = workbook["Angulo"]

    sheet4 = workbook.create_sheet("Torque") # insertando tabla por numero de posicion de izquierda a derecha
    sheet4 = workbook["Torque"]

    sheet5 = workbook.create_sheet("Estadistica") # insertando tabla por numero de posicion de izquierda a derecha
    sheet5 = workbook["Estadistica"]

    sheet['A1'] = '_____'


    # ws = workbook.create_sheet("prueba") # insertando tabla por numero de posicion de izquierda a derecha
    # ws = workbook["prueba"]
    #     # Agregar datos y fórmulas a algunas celdas
    # ws['A1'] = 10
    # ws['B1'] = 20
    # ws['C1'] = "=A1 + A2"  # Esta celda tendrá la fórmula de suma

    # # Evaluar las fórmulas para obtener los valores calculados
    # #ws['A3'].value = ws['A3'].calculate_value()

    # # Crear un gráfico utilizando los datos
    # chart = BarChart()
    # data = Reference(ws, min_col=1, min_row=1, max_col=3, max_row=1)
    # chart.add_data(data)
    # ws.add_chart(chart, "C1")




    #print(arreglo)
    for j in arreglo:
        sheet.append(j)

    for k in arreglo_intt:
        sheet2.append(k)

    for k in arreglo_ang:
        sheet3.append(k)
    
    for k in arreglo_tor:
        sheet4.append(k)

    sheet2.insert_rows(0,1)
    sheet3.insert_rows(0,1)
    sheet4.insert_rows(0,1)

    repeticiones = {}
    # Itera sobre los datos y asigna cada dato a una celda en la fila específica
    for index, dato in enumerate(cajas, start=3):
        if dato in repeticiones:
            repeticiones[dato] += 1
        else:
            repeticiones[dato] = 1 
        sheet2.cell(row=1, column=index).value = dato
        sheet3.cell(row=1, column=index).value = dato
        sheet4.cell(row=1, column=index).value = dato
    #print(repeticiones.keys())
    alineacion_izquierda = Alignment(horizontal='left')
    
    sheet.insert_cols(idx=6,amount=1)
    sheet['F2'] = 'Duracion (dias, horas, minutos, segundos)'
    sheet['H2'] = 'Apriete-Torque'
    
    sheet['I2'] = 'Tiempo Ciclo'
    sheet['J2'] = 'Tiempo Ciclo Horas'
    
    sheet.title = 'Historial'
        # Agrega valores a la nueva columna (no necesitas calcular la diferencia en Python)
    for i in range(3, sheet.max_row + 1):
        formula = f'= IFERROR(ABS(D{i}-E{i}), 0)'  # Suponiendo que "Inicio" está en la columna D y "Fin" en la columna E
        
        # Suponiendo que "Inicio" está en la columna D y "Fin" en la columna E
        formulaB = f'= CONCATENATE( (PRODUCT( DAY(F{i}),24) + HOUR(F{i}) ), ":", MINUTE(F{i}), ":", SECOND(F{i}))'  
        
        formulaC = f'=IFERROR(ABS($D{i}-$E{i}), 0)'  #Tiempo Muerto en Decimal de Horas
        formulaD = f'=IFERROR(ABS($D{i}-$E{i}) * 24, 0)'  #Tiempo Muerto en Decimal de Horas

        sheet.cell(row=i, column=6, value=formula)


        sheet.cell(row=i, column=9, value=formulaC)
        sheet.cell(row=i, column=10, value=formulaD)

    lastfila = get_column_letter(sheet.max_column)+str(sheet.max_row) 
    lastfila2 = get_column_letter(sheet2.max_column)+str(sheet2.max_row) 
    lastfila3 = get_column_letter(sheet3.max_column)+str(sheet3.max_row) 
    lastfila4 = get_column_letter(sheet4.max_column)+str(sheet4.max_row) 


# ####FORMATOS
    formato_hora = NamedStyle(name = 'formato_hora')
    formato_hora.number_format = 'hh:mm:ss'

    # Supongamos que deseas aplicar el formato a la columna A (por ejemplo, de la fila 2 a la fila 100)
    columna = sheet['F']
    for celda in columna[2:sheet.max_row]:  # Excluye la primera fila si tiene encabezados
             celda.number_format = 'dd hh:mm:ss'

    # Supongamos que deseas aplicar el formato a la columna A (por ejemplo, de la fila 2 a la fila 100)
    columna = sheet['J']
    for celda in columna[2:sheet.max_row]:  # Excluye la primera fila si tiene encabezados
            celda.number_format = 'dd hh:mm:ss'

# Iterar sobre todas las columnas y ajustar sus anchos
    for column in sheet2.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                cell.alignment = alineacion_izquierda
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.1
        sheet2.column_dimensions[column_letter].width = adjusted_width
        sheet3.column_dimensions[column_letter].width = adjusted_width
        sheet4.column_dimensions[column_letter].width = adjusted_width
    sheet2.column_dimensions['A'].width = 9
    sheet3.column_dimensions['A'].width = 9
    sheet4.column_dimensions['A'].width = 9

    start_merge = 3
    end_merge = 2
    for tor in repeticiones.keys():
        
        if repeticiones[tor] > 1:

            #print(repeticiones[tor])
            end_merge += repeticiones[tor]
            #print("Start",start_merge)
            #print("End Merge",end_merge)
            sheet2.merge_cells(start_row=1, start_column=start_merge, end_row=1, end_column=end_merge)
            sheet3.merge_cells(start_row=1, start_column=start_merge, end_row=1, end_column=end_merge)
            sheet4.merge_cells(start_row=1, start_column=start_merge, end_row=1, end_column=end_merge)
            start_merge += repeticiones[tor]
        else:
            start_merge += 1
            end_merge += 1
        # Fusionar las celdas


# Iterar sobre todas las columnas y ajustar sus anchos
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                cell.alignment = alineacion_izquierda
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.1
        sheet.column_dimensions[column_letter].width = adjusted_width
    # sheet.column_dimensions['C'].width = 40
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['F'].width = 20
    # sheet.column_dimensions['G'].width = 20
    sheet.column_dimensions['K'].width = 20
    sheet.column_dimensions['E'].width = 19
    sheet.column_dimensions['J'].width = 15
    sheet.column_dimensions['N'].width = 13
    sheet.column_dimensions['O'].width = 15
    sheet.column_dimensions['P'].width = 15
    # Agregar título de la tabla de clientes
    sheet['A1'] = f'Fujikura Automotive México Piedras Negras "Torque"'
    sheet.merge_cells('A1:D1')


    tab = Table(displayName="Table1", ref="A2:" + lastfila)
    tab2 = Table(displayName="Intentos_t", ref="A2:" + lastfila2)
    tab3 = Table(displayName="Angulo", ref="A2:" + lastfila3)
    tab4 = Table(displayName="Torque", ref="A2:" + lastfila4)

    # Agregando Estilos de tabla
    style = TableStyleInfo(name="TableStyleMedium6", showFirstColumn=False,
                    showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tab.tableStyleInfo = style

    # Agregando Estilos de tabla de Inentos Torque
    style2 = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                    showLastColumn=False, showRowStripes=True, showColumnStripes=True)

    # Agregando Estilos de tabla de Inentos Torque
    style3 = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                    showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        # Agregando Estilos de tabla de Inentos Torque
    style4 = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                    showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tab2.tableStyleInfo = style2
    tab3.tableStyleInfo = style3
    tab4.tableStyleInfo = style4


    sheet.add_table(tab)
    sheet2.add_table(tab2)
    sheet3.add_table(tab3)
    sheet4.add_table(tab4)

    # Formulario para calcular por columnas diferentes tareas
    sheet['N2'] = 'Promedio'
    sheet['N3'] = "= TEXT(AVERAGE(F3:F"+str(sheet.max_row)+'), "hh:mm:ss")'
    sheet['O2'] = "No terminados"
    sheet['O3'] = '= COUNTIF(H3:H'+str(sheet.max_row)+',"="&"RESET")'


    tab2 = Table(displayName="Table2", ref="N2:O3")

    # Agregando Formato a la tabla
    formulas = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                    showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tab2.tableStyleInfo = formulas

    sheet.add_table(tab2)



    # Establecer estilos de fuente y color
    first_table_font = Font(color="124B43")  # Azul Marino
    second_table_font = Font(color="0043BB")  # Un tono más claro de rojo
    #Agregando titulos 
    lastRow= str(sheet2.max_row)
    relleno = PatternFill("solid", start_color="5cb800")## Relleno de celdas de color verde
    # sheet5['B2:F2'].fill = "red"
    #PDC-P
    sheet5["B3"]="E1-P"
    sheet5["B4"]=f'=SUM(IntentosT!C3:C{lastRow})'


    #PDC-D
    sheet5["C3"]="E1-D"   
    sheet5["C4"]=f'=SUM(IntentosT!D3:D{lastRow})'

    #PDC-R
    sheet5["D3"]="E1-R"   
    sheet5["D4"]=f"=SUM(IntentosT!E3:E{lastRow})"

    #PDCR-MID
    sheet5["E3"]="E1-RM"
    sheet5["E4"]=f"=SUM(IntentosT!F3:F{lastRow})"

    #PDC-RS
    sheet5["F3"]="E1-RS"
    sheet5["F4"]=f"=SUM(IntentosT!G3:G{lastRow})"


    #MFB-P1
    sheet5["B6"]="A41"
    sheet5["B7"]=f"=SUM(IntentosT!H3:H{lastRow})"

    sheet5["C6"]="A42"
    sheet5["C7"]=f"=SUM(IntentosT!I3:I{lastRow})"

    sheet5["D6"]="A43"
    sheet5["D7"]=f"=SUM(IntentosT!J3:J{lastRow})"

    sheet5["E6"]="A44"
    sheet5["E7"]=f"=SUM(IntentosT!K3:K{lastRow})"

    sheet5["F6"]="A45"
    sheet5["F7"]=f"=SUM(IntentosT!L3:L{lastRow})"

    sheet5["G6"]="A46"
    sheet5["G7"]=f"=SUM(IntentosT!M3:M{lastRow})"

    sheet5["H6"]="A47"
    sheet5["H7"]=f"=SUM(IntentosT!N3:N{lastRow})"
    
    #MFB-S
    sheet5["B9"]="A51"
    sheet5["B10"]=f"=SUM(IntentosT!O3:O{lastRow})"
    
    sheet5["C9"]="A52"
    sheet5["C10"]=f"=SUM(IntentosT!P3:P{lastRow})"

    sheet5["D9"]="A53"
    sheet5["D10"]=f"=SUM(IntentosT!Q3:Q{lastRow})"
    sheet5["E9"]="A54"
    sheet5["E10"]=f"=SUM(IntentosT!R3:R{lastRow})"
    sheet5["F9"]="A55"
    sheet5["F10"]=f"=SUM(IntentosT!S3:S{lastRow})"
    sheet5["G9"]="A56"
    sheet5["G10"]=f"=SUM(IntentosT!T3:T{lastRow})"

    #MFB-E
    sheet5["B12"]="E1"
    sheet5["B13"]=f"=SUM(IntentosT!U3:U{lastRow})"
    sheet5["C12"]="A1"
    sheet5["C13"]=f"=SUM(IntentosT!V3:V{lastRow})"
    sheet5["D12"]="A2"
    sheet5["D13"]=f"=SUM(IntentosT!W3:W{lastRow})"





    #MPFB-P2
    sheet5["B15"]="A20"
    sheet5["B16"]=f"=SUM(IntentosT!X3:X{lastRow})"

    sheet5["C15"]="A21"
    sheet5["C16"]=f"=SUM(IntentosT!Y3:Y{lastRow})"

    sheet5["D15"]="A22"
    sheet5["D16"]=f"=SUM(IntentosT!Z3:Z{lastRow})"

    sheet5["E15"]="A23"
    sheet5["E16"]=f"=SUM(IntentosT!AA3:AA{lastRow})"

    sheet5["F15"]="A24"
    sheet5["F16"]=f"=SUM(IntentosT!AB3:AB{lastRow})"

    sheet5["G15"]="A25"
    sheet5["G16"]=f"=SUM(IntentosT!AC3:AC{lastRow})"

    sheet5["H15"]="A26"
    sheet5["H16"]=f"=SUM(IntentosT!AD3:AD{lastRow})"

    sheet5["I15"]="A27"
    sheet5["I16"]=f"=SUM(IntentosT!AE3:AE{lastRow})"

    sheet5["J15"]="A28"
    sheet5["J16"]=f"=SUM(IntentosT!AF3:AF{lastRow})"

    sheet5["K15"]="A29"
    sheet5["K16"]=f"=SUM(IntentosT!AG3:AG{lastRow})"

    sheet5["L15"]="A30"
    sheet5["L16"]=f"=SUM(IntentosT!AH3:AH{lastRow})"



    #Battery
    sheet5["B18"]="BT"
    sheet5["B19"]=f"=SUM(IntentosT!AI3:AI{lastRow})"
    #Battery 2
    sheet5["C18"]="BT-2"
    sheet5["C19"]=f"=SUM(IntentosT!AJ3:AJ{lastRow})"

    # Obtener los datos de las celdas que quieres graficar
    valores = []
    categorias = []
    chart1 = BarChart()

    #PDC'S
    data =  Reference(sheet5, min_col=2, min_row=3, max_row=4, max_col=6)
    chart1.add_data(data,titles_from_data=True)

    #MFB-P1
    data =  Reference(sheet5, min_col=2, min_row=6, max_row=7, max_col=8)
    cats =  Reference(sheet5, min_col=1, min_row=1, max_row=1, max_col=1)
    chart1.add_data(data,titles_from_data=True)
    chart1.set_categories(cats)

    #MFB-S
    data =  Reference(sheet5, min_col=2, min_row=9, max_row=10, max_col=7)
    chart1.add_data(data,titles_from_data=True)
    #chart1.set_categories(cats)

    #MFB-E
    data =  Reference(sheet5, min_col=2, min_row=12, max_row=13, max_col=4)
    chart1.add_data(data,titles_from_data=True)
    #chart1.set_categories(cats)

    #MFB-P2
    data =  Reference(sheet5, min_col=2, min_row=15, max_row=16, max_col=12)
    chart1.add_data(data,titles_from_data=True)
    #chart1.set_categories(cats)


    #Battery's
    data =  Reference(sheet5, min_col=2, min_row=18, max_row=19, max_col=3)
    chart1.add_data(data,titles_from_data=True)
    

    chart1.title = 'Reintentos Torque'
    chart1.height = 7.5 # default is 7.5
    chart1.width = 25 # default is 15
    # data = valores
    # cats = categorias
    sheet5.add_chart(chart1, "B21")
    sheet5["C17"]="BATTERY-2"
    sheet5["C17"].fill=relleno

    sheet5["B17"]="BATTERY"
    sheet5["B17"].fill=relleno

    sheet5["B14"]="MFB-P2"
    sheet5["B14"].fill=relleno
    rango_celdas = 'B14:L14'
    sheet5.merge_cells(rango_celdas)

    sheet5["B11"]="MFB-E"
    sheet5["B11"].fill=relleno
    rango_celdas = 'B11:D11'
    sheet5.merge_cells(rango_celdas)

    sheet5["B8"]="MFB-S"
    sheet5["B8"].fill=relleno
    rango_celdas = 'B8:G8'
    sheet5.merge_cells(rango_celdas)

    sheet5["B5"]="MFB-P1"
    sheet5["B5"].fill=relleno
    # Definimos el rango de celdas donde queremos aplicar el color
    rango_celdas = 'B5:H5'
    sheet5.merge_cells(rango_celdas)

    sheet5["F2"]="PDC-RS"
    sheet5["F2"].fill = relleno

    sheet5["E2"]="PDC-RMID"
    sheet5["E2"].fill = relleno

    sheet5["D2"]="PDC-R"   
    sheet5["D2"].fill = relleno

    sheet5["C2"]="PDC-D"   
    sheet5["C2"].fill = relleno

    sheet5["B2"]="PDC-P"
    sheet5["B2"].fill = relleno

    # Guardar el libro de Excel en un objeto en memoria
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    # Enviar el archivo como respuesta para descarga
    return send_file(
        output,
        as_attachment=True,
        download_name='Fujikura Automotive México Piedras Negras.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@app.route('/descargarTorque/<tool>/<val1>/<val2>')
def descargarTorque(tool, val1, val2):
    query = '''
    SELECT 
    t.*, 
    ROUND(JSON_EXTRACT(REGISTRO, '$.angle'), 2) as round_angle,
    ROUND(JSON_EXTRACT(REGISTRO, '$.torque'), 2) as round_torque,
    (ROUND(JSON_EXTRACT(REGISTRO, '$.angle'), 2) + ROUND(JSON_EXTRACT(REGISTRO, '$.torque'), 2)) as SUMA_angle_torque,
    h.HM AS HM_pertenece,
    
    JSON_EXTRACT(h.ANGULO, '$.MFB-P2.A20') AS A20_angulo,
    JSON_EXTRACT(h.TORQUE, '$.MFB-P2.A20') AS A20_torque,
    ( JSON_EXTRACT(h.ANGULO, '$.MFB-P2.A20') + JSON_EXTRACT(h.TORQUE, '$.MFB-P2.A20') ) as SUMA_A20,
    JSON_EXTRACT(h.INTENTOS_T, '$.MFB-P2.A20') AS A20_intentos,
    
    
    JSON_EXTRACT(h.ANGULO, '$.MFB-P2.A25') AS A25_angulo,
    JSON_EXTRACT(h.TORQUE, '$.MFB-P2.A25') AS A25_torque,
    ( JSON_EXTRACT(h.ANGULO, '$.MFB-P2.A25') + JSON_EXTRACT(h.TORQUE, '$.MFB-P2.A25') ) as SUMA_A25,
    JSON_EXTRACT(h.INTENTOS_T, '$.MFB-P2.A25') AS A25_intentos,
    
    JSON_EXTRACT(h.ANGULO, '$.MFB-P2.A30') AS A30_angulo,
    JSON_EXTRACT(h.TORQUE, '$.MFB-P2.A30') AS A30_torque,
    ( JSON_EXTRACT(h.ANGULO, '$.MFB-P2.A30') + JSON_EXTRACT(h.TORQUE, '$.MFB-P2.A30') ) as SUMA_A30,
    JSON_EXTRACT(h.INTENTOS_T, '$.MFB-P2.A30') AS A30_intentos,
    
    JSON_EXTRACT(h.ANGULO, '$.PDC-RS.E1') AS PDCRS_angulo,
    JSON_EXTRACT(h.TORQUE, '$.PDC-RS.E1') AS PDCRS_torque,
    ( JSON_EXTRACT(h.ANGULO, '$.PDC-RS.E1') + JSON_EXTRACT(h.TORQUE, '$.PDC-RS.E1') ) as SUMA_PDCRS,
    JSON_EXTRACT(h.INTENTOS_T, '$.PDC-RS.E1') AS PDCRS_intentos,
    
    JSON_EXTRACT(h.ANGULO, '$.PDC-RMID.E1') AS PDCRMID_angulo,
    JSON_EXTRACT(h.TORQUE, '$.PDC-RMID.E1') AS PDCRMID_torque,
    ( JSON_EXTRACT(h.ANGULO, '$.PDC-RMID.E1') + JSON_EXTRACT(h.TORQUE, '$.PDC-RMID.E1') ) as SUMA_PDCRMID,
    JSON_EXTRACT(h.INTENTOS_T, '$.PDC-RMID.E1') AS PDCRMID_intentos,
    
    JSON_EXTRACT(h.ANGULO, '$.PDC-R.E1') AS PDCR_angulo,
    JSON_EXTRACT(h.TORQUE, '$.PDC-R.E1') AS PDCR_torque,
    ( JSON_EXTRACT(h.ANGULO, '$.PDC-R.E1') + JSON_EXTRACT(h.TORQUE, '$.PDC-R.E1') ) as SUMA_PDCR,
    JSON_EXTRACT(h.INTENTOS_T, '$.PDC-R.E1') AS PDCR_intentos,
    
    JSON_EXTRACT(h.ANGULO, '$.MFB-P1.A41') AS A41_angulo,
    JSON_EXTRACT(h.TORQUE, '$.MFB-P1.A41') AS A41_torque,
    ( JSON_EXTRACT(h.ANGULO, '$.MFB-P1.A41') + JSON_EXTRACT(h.TORQUE, '$.MFB-P1.A41') ) as SUMA_A41,
    JSON_EXTRACT(h.INTENTOS_T, '$.MFB-P1.A41') AS A41_intentos,
    
    JSON_EXTRACT(h.ANGULO, '$.MFB-P1.A46') AS A46_angulo,
    JSON_EXTRACT(h.TORQUE, '$.MFB-P1.A46') AS A46_torque,
    ( JSON_EXTRACT(h.ANGULO, '$.MFB-P1.A46') + JSON_EXTRACT(h.TORQUE, '$.MFB-P1.A46') ) as SUMA_A46,
    JSON_EXTRACT(h.INTENTOS_T, '$.MFB-P1.A46') AS A46_intentos,
    
    JSON_EXTRACT(h.ANGULO, '$.MFB-P1.A47') AS A47_angulo,
    JSON_EXTRACT(h.TORQUE, '$.MFB-P1.A47') AS A47_torque,
    ( JSON_EXTRACT(h.ANGULO, '$.MFB-P1.A47') + JSON_EXTRACT(h.TORQUE, '$.MFB-P1.A47') ) as SUMA_A47,
    JSON_EXTRACT(h.INTENTOS_T, '$.MFB-P1.A47') AS A46_intentos,
    
    JSON_EXTRACT(h.ANGULO, '$.MFB-E.A1') AS AMG_A1_angulo,
    JSON_EXTRACT(h.TORQUE, '$.MFB-E.A1') AS AMG_A1_torque,
    ( JSON_EXTRACT(h.ANGULO, '$.MFB-E.A1') + JSON_EXTRACT(h.TORQUE, '$.MFB-E.A1') ) as SUMA_AMG_A1,
    JSON_EXTRACT(h.INTENTOS_T, '$.MFB-E.A1') AS AMG_A1_intentos,
    
    JSON_EXTRACT(h.ANGULO, '$.MFB-E.A2') AS AMG_A2_angulo,
    JSON_EXTRACT(h.TORQUE, '$.MFB-E.A2') AS AMG_A2_torque,
    ( JSON_EXTRACT(h.ANGULO, '$.MFB-E.A2') + JSON_EXTRACT(h.TORQUE, '$.MFB-E.A2') ) as SUMA_AMG_A2,
    JSON_EXTRACT(h.INTENTOS_T, '$.MFB-E.A2') AS AMG_A2_intentos,
    
    JSON_EXTRACT(h.ANGULO, '$.MFB-E.E1') AS AMG_E1_angulo,
    JSON_EXTRACT(h.TORQUE, '$.MFB-E.E1') AS AMG_E1_torque,
    ( JSON_EXTRACT(h.ANGULO, '$.MFB-E.E1') + JSON_EXTRACT(h.TORQUE, '$.MFB-E.E1') ) as SUMA_AMG_E1,
    JSON_EXTRACT(h.INTENTOS_T, '$.MFB-E.E1') AS AMG_E1_intentos,
    
    JSON_EXTRACT(h.ANGULO, '$.MFB-S.A51') AS A51_angulo,
    JSON_EXTRACT(h.TORQUE, '$.MFB-S.A51') AS A51_torque,
    ( JSON_EXTRACT(h.ANGULO, '$.MFB-S.A51') + JSON_EXTRACT(h.TORQUE, '$.MFB-S.A51') ) as SUMA_A51,
    JSON_EXTRACT(h.INTENTOS_T, '$.MFB-S.A51') AS A51_intentos,
    
    JSON_EXTRACT(h.ANGULO, '$.MFB-S.A52') AS A52_angulo,
    JSON_EXTRACT(h.TORQUE, '$.MFB-S.A52') AS A52_torque,
    ( JSON_EXTRACT(h.ANGULO, '$.MFB-S.A52') + JSON_EXTRACT(h.TORQUE, '$.MFB-S.A52') ) as SUMA_A52,
    JSON_EXTRACT(h.INTENTOS_T, '$.MFB-S.A52') AS A52_intentos  
'''

    # if val1.includes("HM"):
    #     formWhere = query + f'''
    # FROM 
    #     {database}.torque_info t
    # INNER JOIN 
    #     {database}.historial h ON t.FECHA BETWEEN h.INICIO AND h.FIN

    # WHERE h.HM = "{val1}" 
    # AND t.HERRAMIENTA = "{tool}"
    # ORDER BY t.ID DESC;
    # '''
    # else:
    formWhere = query + f'''
        FROM 
            {database}.torque_info t
        INNER JOIN 
            {database}.historial h ON t.FECHA BETWEEN h.INICIO AND h.FIN

        WHERE t.FECHA >= "{val1}" AND t.FECHA <= "{val2}"
        AND t.HERRAMIENTA = "{tool}"
        
        ''' 
    
    try:
        connection = pymysql.connect(host = host, user = user, passwd = password, database = database, cursorclass=pymysql.cursors.DictCursor)

    except Exception as ex:
        print("myJsonResponse connection Exception: ", ex)
        return {"exception": ex.args}
    try:
        with connection.cursor() as cursor:
                items = cursor.execute(formWhere)
                result = cursor.fetchall()
                print(formWhere)
                print(result)
                h =  result[0]
                r = json.loads(result[0]["REGISTRO"])
                rKeys = list(r.keys())
                #print(result[0]["REGISTRO"])
                print(rKeys)
                arreglo = []
                valores_fila = ','.join( str(valor) for valor in h)
                li = list(valores_fila.split(","))
                print(li)
                
                # Encontrar la ubicación de "REGISTRO"
                ubicacion_registro = li.index('REGISTRO')

                #li.remove('REGISTRO')

                # Insertar los valores en la lista después de "REGISTRO"
                li[ubicacion_registro + 1:ubicacion_registro + 1] = rKeys
                
                li.remove("REGISTRO")
                print(li)


                arreglo.append(li)
                for fila in result:
                        dato= []
                        arreglo.append(dato)
                        for i in fila:  
                                if "REGISTRO" in i:

                                    registro = json.loads(fila[i])
                                    lista_de_registros = list(registro.values())
                                    dato.extend(lista_de_registros)

                                elif not "REGISTRO" in i:
                                    dato.append(fila[i])


                                # Define las dos fechas como cadenas de texto
                                # Convierte las cadenas de texto en objetos datetime
                                #fecha1 = datetime.strptime(fecha1_str, "%d-%m-%Y %H:%M:%S")
                                #fecha2 = datetime.strptime(fecha2_str, "%d-%m-%Y %H:%M:%S")

                #print(arreglo)
                workbook = Workbook()
                sheet = workbook.active
                sheet.title = 'Torque_info'
                for j in arreglo:
                    sheet.append(j)
                
                    # Guardar el libro de Excel en un objeto en memoria
                output = io.BytesIO()
                workbook.save(output)
                output.seek(0)

                # Enviar el archivo como respuesta para descarga
                return send_file(
                    output,
                    as_attachment=True,
                    download_name='Fujikura Automotive México Piedras Negras.xlsx',
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )

    except Exception as ex:
        print("myJsonResponse connection Exception: ", ex)
        return {"exception": ex.args}